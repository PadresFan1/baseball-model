import sys
sys.stdout.reconfigure(encoding='utf-8')

class _Tee:
    """Write to both the real stdout and a capture buffer simultaneously."""
    def __init__(self):
        self._real = sys.stdout
        self._buf  = __import__('io').StringIO()
    def write(self, msg):
        self._real.write(msg)
        self._buf.write(msg)
    def flush(self):
        self._real.flush()
    def getvalue(self):
        return self._buf.getvalue()

_tee = _Tee()
sys.stdout = _tee

import pandas as pd
import math
import requests
import statsapi
import json
import os
from dotenv import load_dotenv
load_dotenv(dotenv_path=r'C:\Users\super\baseball-model\.env')
from datetime import date
import numpy as np
from io import StringIO

# Team Abbreviation Translation - MLB API to Fangraphs
TEAM_MAP = {
    'ATH': 'ATH',
    'PIT': 'PIT',
    'SD': 'SDP',
    'SEA': 'SEA',
    'SF': 'SFG',
    'STL': 'STL',
    'TB': 'TBR',
    'TEX': 'TEX',
    'TOR': 'TOR',
    'MIN': 'MIN',
    'PHI': 'PHI',
    'ATL': 'ATL',
    'CWS': 'CHW',
    'MIA': 'MIA',
    'NYY': 'NYY',
    'MIL': 'MIL',
    'LAA': 'LAA',
    'AZ': 'ARI',
    'BAL': 'BAL',
    'BOS': 'BOS',
    'CHC': 'CHC',
    'CIN': 'CIN',
    'CLE': 'CLE',
    'COL': 'COL',
    'DET': 'DET',
    'HOU': 'HOU',
    'KC': 'KCR',
    'LAD': 'LAD',
    'WSH': 'WSN',
    'NYM': 'NYM'
}

NAME_TO_FG = {
    'Athletics': 'ATH',
    'Pittsburgh Pirates': 'PIT',
    'San Diego Padres': 'SDP',
    'Seattle Mariners': 'SEA',
    'San Francisco Giants': 'SFG',
    'St. Louis Cardinals': 'STL',
    'Tampa Bay Rays': 'TBR',
    'Texas Rangers': 'TEX',
    'Toronto Blue Jays': 'TOR',
    'Minnesota Twins': 'MIN',
    'Philadelphia Phillies': 'PHI',
    'Atlanta Braves': 'ATL',
    'Chicago White Sox': 'CHW',
    'Miami Marlins': 'MIA',
    'New York Yankees': 'NYY',
    'Milwaukee Brewers': 'MIL',
    'Los Angeles Angels': 'LAA',
    'Arizona Diamondbacks': 'ARI',
    'Baltimore Orioles': 'BAL',
    'Boston Red Sox': 'BOS',
    'Chicago Cubs': 'CHC',
    'Cincinnati Reds': 'CIN',
    'Cleveland Guardians': 'CLE',
    'Colorado Rockies': 'COL',
    'Detroit Tigers': 'DET',
    'Houston Astros': 'HOU',
    'Kansas City Royals': 'KCR',
    'Los Angeles Dodgers': 'LAD',
    'Washington Nationals': 'WSN',
    'New York Mets': 'NYM'
}

PARK_FACTOR_TEAM_MAP = {
    'Angels': 'LAA', 'Orioles': 'BAL', 'Red Sox': 'BOS',
    'White Sox': 'CHW', 'Guardians': 'CLE', 'Tigers': 'DET',
    'Royals': 'KCR', 'Twins': 'MIN', 'Yankees': 'NYY',
    'Athletics': 'ATH', 'Mariners': 'SEA', 'Rays': 'TBR',
    'Rangers': 'TEX', 'Blue Jays': 'TOR', 'Diamondbacks': 'ARI',
    'Braves': 'ATL', 'Cubs': 'CHC', 'Reds': 'CIN',
    'Rockies': 'COL', 'Dodgers': 'LAD', 'Marlins': 'MIA',
    'Brewers': 'MIL', 'Mets': 'NYM', 'Phillies': 'PHI',
    'Pirates': 'PIT', 'Cardinals': 'STL', 'Padres': 'SDP',
    'Giants': 'SFG', 'Nationals': 'WSN', 'Astros': 'HOU'
}

from datetime import timezone, timedelta, datetime

API_KEY  = os.getenv('API_KEY')
BANKROLL = float(os.getenv('BANKROLL', '1000'))
KELLY_FRACTION = 0.5          # half-Kelly reduces variance vs full Kelly
MAX_BANKROLL_EXPOSURE = 0.15  # hard cap per bet regardless of Kelly edge
OU_BETTING_ENABLED = False    # O/U shows negative ROI in backtesting — disabled until calibrated
print(f"API Key loaded: {API_KEY is not None}")

from zoneinfo import ZoneInfo as _ZoneInfo
MT = _ZoneInfo('America/Denver')  # handles MST/MDT automatically

# Starter/bullpen split
STARTER_INNINGS = 5
TOTAL_INNINGS = 9
STARTER_WEIGHT = STARTER_INNINGS / TOTAL_INNINGS
BULLPEN_WEIGHT = 1 - STARTER_WEIGHT
MIN_INNINGS = 15
MAX_RA9 = 7.00

# Weights derived from round-5 backtest optimization (top-10 trials, n>=300)
# Offense: k% and bb% removed — backtest showed no positive contribution
OFFENSE_WEIGHTS = {
    'rolling': 0.090,
    'woba':    0.378,
    'xwoba':   0.264,
    'babip':   0.268,
}

# Pitching: normalized over 5 available stats (xwoba_pit/babip_pit/barrel_pit not yet in live model)
PITCHING_WEIGHTS = {
    'rolling': 0.078,
    'fip':     0.240,
    'xfip':    0.220,
    'k_pct':   0.192,
    'bb_pct':  0.270,
}

NUM_SIMULATIONS = 5000

# ── Log5 inference layer ──────────────────────────────────────────────────────
import pickle as _pickle

_LOG5_PA_WEIGHTS = np.array([0.123, 0.120, 0.117, 0.114, 0.111, 0.108, 0.105, 0.102, 0.100])
_LOG5_FEATURES   = ['d_log5_woba', 'd_log5_xwoba', 'd_xfip']  # player-only; no market anchor

def _log5_matchup(B, P, L):
    B = float(np.clip(B, 0.001, 0.999))
    P = float(np.clip(P, 0.001, 0.999))
    L = float(np.clip(L, 0.001, 0.999))
    num = (B * P) / L
    den = num + ((1.0 - B) * (1.0 - P)) / (1.0 - L)
    return num / den if den > 0 else L

def _aggregate_log5(matchup_dict, lineup, fallback):
    slots   = lineup[:9]
    metrics = np.array([matchup_dict.get(pid, fallback) for pid in slots], dtype=float)
    if len(metrics) < 9:
        metrics = np.pad(metrics, (0, 9 - len(metrics)), constant_values=fallback)
    return float(np.dot(metrics, _LOG5_PA_WEIGHTS))

def load_log5_assets():
    """Load trained regression model and player snapshot from disk. Returns (bundle, snapshot) or (None, None)."""
    try:
        with open('models/log5_regression.pkl', 'rb') as f:
            bundle = _pickle.load(f)
        with open('historical_data/player_snapshot.json', 'r') as f:
            snapshot = json.load(f)
        as_of = snapshot.get('as_of', 'unknown')
        n_bat = snapshot.get('n_batters', 0)
        n_pit = snapshot.get('n_pitchers', 0)
        print(f"[Log5] Model loaded (C={bundle['C']}, trained on {bundle['n_train']:,} games)")
        print(f"[Log5] Player snapshot: {n_bat:,} batters, {n_pit:,} pitchers, as_of {as_of}")
        return bundle, snapshot
    except FileNotFoundError:
        print("[Log5] Production assets not found — run backtest.py with RUN_MODE='build_production_model' first. Falling back to Poisson.")
        return None, None
    except Exception as e:
        print(f"[Log5] Failed to load assets: {e}. Falling back to Poisson.")
        return None, None

_lineup_history_cache = None  # loaded once per run; shared across all build_projected_lineup calls

def build_projected_lineup(team_fg, n_games=10):
    """
    Returns a 9-slot projected batting order for team_fg based on the most
    frequently used player at each slot across the last n_games confirmed
    lineups from the historical lineup cache.

    Slots with no data stay as 0 — compute_log5_win_prob treats those as
    league-average batters via its fallback path.
    """
    global _lineup_history_cache
    if _lineup_history_cache is None:
        try:
            with open('historical_data/game_lineups.json') as _f:
                _lineup_history_cache = json.load(_f)
        except Exception:
            _lineup_history_cache = {}

    team_games = []
    for season_data in _lineup_history_cache.values():
        for gdata in season_data.values():
            if gdata.get('home_fg') == team_fg:
                raw = gdata.get('home_lineup', [])
            elif gdata.get('away_fg') == team_fg:
                raw = gdata.get('away_lineup', [])
            else:
                continue
            if raw and len(raw) >= 9 and gdata.get('date'):
                team_games.append((gdata['date'], [int(p) for p in raw[:9]]))

    if not team_games:
        return []

    team_games.sort(key=lambda x: x[0], reverse=True)
    recent = [lu for _, lu in team_games[:n_games]]

    from collections import Counter
    projected = []
    for slot in range(9):
        votes = Counter(lu[slot] for lu in recent if slot < len(lu))
        projected.append(votes.most_common(1)[0][0] if votes else 0)
    return projected


def compute_log5_win_prob(home_lineup, away_lineup,
                          home_pitcher_id, away_pitcher_id,
                          home_pitcher_hand, away_pitcher_hand,
                          log5_bundle, player_snapshot):
    """
    Compute P(home win) from the player-only Log5 logistic regression.
    Features: d_log5_woba, d_log5_xwoba, d_xfip — no market anchor.
    Returns (home_prob, away_prob) or None if data is insufficient.

    home/away_pitcher_hand: 'L' or 'R' — the hand the opposing lineup faces.
    """
    try:
        batters   = player_snapshot['batters']
        pitchers  = player_snapshot['pitchers']
        lg        = player_snapshot['lg_avgs']

        # League averages as fallbacks
        lg_woba_h  = lg.get(away_pitcher_hand, {}).get('woba',  0.320)
        lg_xwoba_h = lg.get(away_pitcher_hand, {}).get('xwoba', 0.315)
        lg_woba_a  = lg.get(home_pitcher_hand, {}).get('woba',  0.320)
        lg_xwoba_a = lg.get(home_pitcher_hand, {}).get('xwoba', 0.315)
        lg_xfip    = lg.get('pit', {}).get('xfip',      4.10)
        lg_xwoba_p = lg.get('pit', {}).get('xwoba_pit', 0.315)

        # Pitcher xwOBA allowed (P in Log5 formula)
        away_pit  = pitchers.get(str(away_pitcher_id), {})
        home_pit  = pitchers.get(str(home_pitcher_id), {})
        away_xwoba_pit = away_pit.get('xwoba_pit') or lg_xwoba_p
        home_xwoba_pit = home_pit.get('xwoba_pit') or lg_xwoba_p
        away_xfip = away_pit.get('xfip') or lg_xfip
        home_xfip = home_pit.get('xfip') or lg_xfip

        # Home offense vs away pitcher hand
        h_log5_woba  = {}
        h_log5_xwoba = {}
        h_raw_woba   = {}
        h_raw_xwoba  = {}
        for pid in home_lineup:
            entry = batters.get(str(pid), {}).get(away_pitcher_hand, {})
            bw  = entry.get('woba')
            bxw = entry.get('xwoba')
            if bw  is not None:
                h_raw_woba[pid]  = bw
                h_log5_woba[pid] = _log5_matchup(bw,  away_xwoba_pit, lg_woba_h)
            if bxw is not None:
                h_raw_xwoba[pid]  = bxw
                h_log5_xwoba[pid] = _log5_matchup(bxw, away_xwoba_pit, lg_xwoba_h)

        # Away offense vs home pitcher hand
        a_log5_woba  = {}
        a_log5_xwoba = {}
        a_raw_woba   = {}
        a_raw_xwoba  = {}
        for pid in away_lineup:
            entry = batters.get(str(pid), {}).get(home_pitcher_hand, {})
            bw  = entry.get('woba')
            bxw = entry.get('xwoba')
            if bw  is not None:
                a_raw_woba[pid]  = bw
                a_log5_woba[pid] = _log5_matchup(bw,  home_xwoba_pit, lg_woba_a)
            if bxw is not None:
                a_raw_xwoba[pid]  = bxw
                a_log5_xwoba[pid] = _log5_matchup(bxw, home_xwoba_pit, lg_xwoba_a)

        # Fallbacks for missing batters
        h_fb_woba  = _log5_matchup(lg_woba_h,  away_xwoba_pit, lg_woba_h)
        h_fb_xwoba = _log5_matchup(lg_xwoba_h, away_xwoba_pit, lg_xwoba_h)
        a_fb_woba  = _log5_matchup(lg_woba_a,  home_xwoba_pit, lg_woba_a)
        a_fb_xwoba = _log5_matchup(lg_xwoba_a, home_xwoba_pit, lg_xwoba_a)

        home_log5_woba  = _aggregate_log5(h_log5_woba,  home_lineup, h_fb_woba)
        home_log5_xwoba = _aggregate_log5(h_log5_xwoba, home_lineup, h_fb_xwoba)
        away_log5_woba  = _aggregate_log5(a_log5_woba,  away_lineup, a_fb_woba)
        away_log5_xwoba = _aggregate_log5(a_log5_xwoba, away_lineup, a_fb_xwoba)

        home_xfip_r = lg_xfip / home_xfip if home_xfip > 0 else 1.0
        away_xfip_r = lg_xfip / away_xfip if away_xfip > 0 else 1.0

        # Player-only feature vector — all 3 features standardized, no market anchor
        features = np.array([[
            home_log5_woba  - away_log5_woba,
            home_log5_xwoba - away_log5_xwoba,
            home_xfip_r     - away_xfip_r,
        ]], dtype=float)

        mu  = np.array(log5_bundle['mu'])
        sig = np.array(log5_bundle['sig'])
        X_s = (features - mu) / sig

        home_prob = float(log5_bundle['model'].predict_proba(X_s)[0, 1])

        # Apply Platt calibration if present in bundle
        if 'calibrator' in log5_bundle:
            try:
                _raw_lg   = np.log(np.clip(home_prob, 1e-7, 1-1e-7) /
                                   (1 - np.clip(home_prob, 1e-7, 1-1e-7)))
                home_prob = float(log5_bundle['calibrator'].predict_proba(
                    np.array([[_raw_lg]]))[0, 1])
            except Exception:
                pass

        disp_stats = {
            'h_bat_woba':  _aggregate_log5(h_raw_woba,  home_lineup, lg_woba_h),
            'h_bat_xwoba': _aggregate_log5(h_raw_xwoba, home_lineup, lg_xwoba_h),
            'a_bat_woba':  _aggregate_log5(a_raw_woba,  away_lineup, lg_woba_a),
            'a_bat_xwoba': _aggregate_log5(a_raw_xwoba, away_lineup, lg_xwoba_a),
            'h_pit_xfip':  home_xfip,
            'a_pit_xfip':  away_xfip,
            'h_pit_hand':  home_pitcher_hand,
            'a_pit_hand':  away_pitcher_hand,
        }
        return home_prob, 1.0 - home_prob, disp_stats

    except Exception as e:
        return None

# ─────────────────────────────────────────────────────────────────────────────

def load_constants(year=2026):
    woba_fip = pd.read_csv('constants/woba_fip_constants.csv')
    row = woba_fip[woba_fip['Season'] == year].iloc[0]
    
    woba_weights = {
        'wBB': row['wBB'],
        'wHBP': row['wHBP'],
        'w1B': row['w1B'],
        'w2B': row['w2B'],
        'w3B': row['w3B'],
        'wHR': row['wHR']
    }
    fip_constant = row['cFIP']
    
    park_factors = pd.read_csv('constants/park_factors.csv')
    park_factors_hand = pd.read_csv('constants/park_factors_handedness.csv')
    
    return woba_weights, fip_constant, park_factors, park_factors_hand

WOBA_WEIGHTS, FIP_CONSTANT, PARK_FACTORS, PARK_FACTORS_HAND = load_constants(2026)

# Run schedule for predictions log — 5 windows (all times MST)
RUN_SCHEDULE = {
    1: '6PM_Overnight',   # 6:00 PM MST night-before: opening lines drop
    2: '11PM_Overnight',  # 11:00 PM MST night-before: late/west-coast lines
    3: '5AM_Morning',     # 5:00 AM MST day-of: overnight line movement
    4: '11AM_Midday',     # 11:00 AM MST day-of: pre-slate check
    5: '3PM_Afternoon',   # 3:00 PM MST day-of: final before evening games
}

_OVERNIGHT_WINDOWS = {'6PM_Overnight', '11PM_Overnight'}

def _compute_run_context():
    """Return (run_num, run_window, target_date) based on current local time."""
    from datetime import datetime as _dtnow, timedelta
    hour = _dtnow.now().hour
    if 17 <= hour <= 20:        run_num, run_window = 1, '6PM_Overnight'
    elif hour >= 21 or hour <= 2: run_num, run_window = 2, '11PM_Overnight'
    elif 3 <= hour <= 7:        run_num, run_window = 3, '5AM_Morning'
    elif 8 <= hour <= 13:       run_num, run_window = 4, '11AM_Midday'
    else:                       run_num, run_window = 5, '3PM_Afternoon'
    target_date = date.today() + timedelta(days=1) if run_window in _OVERNIGHT_WINDOWS else date.today()
    return run_num, run_window, target_date

def american_to_prob(line):
    if line < 0:
        return abs(line) / (abs(line) + 100)
    else:
        return 100 / (line + 100)

def calculate_kelly_stake(bankroll, market_odds, model_prob, odds_type='american'):
    """
    Returns recommended dollar stake using 1/2 Kelly criterion capped at MAX_BANKROLL_EXPOSURE.
    American odds are converted to decimal; b = decimal_odds - 1 (net odds per $1 wagered).
    fraction = 0.5 * ((b * p - q) / b); returns $0.00 when edge is zero or negative.
    """
    if odds_type == 'american':
        decimal_odds = (1 + market_odds / 100) if market_odds > 0 else (1 + 100 / abs(market_odds))
    else:
        decimal_odds = float(market_odds)
    b = decimal_odds - 1
    if b <= 0:
        return 0.0
    fraction = 0.5 * ((b * model_prob - (1 - model_prob)) / b)
    if fraction <= 0:
        return 0.0
    fraction = min(fraction, MAX_BANKROLL_EXPOSURE)
    return float(math.floor(bankroll * fraction + 0.5))

def kelly_bet_size(model_prob, american_odds, bankroll=None, fraction=KELLY_FRACTION):
    """Returns (kelly_pct, bet_amount) using calculate_kelly_stake as the canonical engine."""
    if bankroll is None:
        bankroll = BANKROLL
    amount = calculate_kelly_stake(bankroll, american_odds, model_prob)
    pct = round(amount / bankroll * 100, 1) if bankroll > 0 else 0.0
    return pct, amount

def ip_to_float(ip_str):
    """Convert MLB innings pitched string (e.g. '5.1'=5⅓, '5.2'=5⅔) to decimal."""
    try:
        s = str(ip_str)
        if '.' in s:
            whole, frac = s.split('.')
            return int(whole) + int(frac) / 3.0
        return float(s)
    except (ValueError, AttributeError):
        return 0.0

def get_cached_odds():
    cache_file='cache/odds_cache.json'

    # Check if cache exists and is fresh
    if os.path.exists(cache_file):
        modified_time = os.path.getmtime(cache_file)
        from datetime import datetime
        age_hours = (datetime.now().timestamp() - modified_time) / 3600
        if age_hours < 1.0:
            print ("Loading odds from cache. . .")
            with open(cache_file,'r') as f:
                return json.load(f)
            
    # Cache is old or doesn't exist - call the API
    print("Fetching fresh odds from API . . .")
    data = get_mlb_odds()

    # Save to cache
    with open(cache_file,'w') as f:
        json.dump(data,f)

    return data

def get_upcoming_games(odds_data, target_date=None):
    now = datetime.now(timezone.utc)
    if target_date is None:
        target_date = now.astimezone(MT).date()
    upcoming = []
    for game in odds_data:
        game_time = datetime.strptime(game['commence_time'], '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc)
        game_time_mst = game_time.astimezone(MT)
        if game_time > now and game_time_mst.date() == target_date:
            upcoming.append(game)
    return upcoming

def get_todays_game_statuses(target_date=None):
    today = (target_date or date.today()).strftime('%Y-%m-%d')
    schedule = statsapi.schedule(date=today)
    schedule = sorted(schedule, key=lambda g: g.get('game_datetime', ''))
    statuses = {}
    pair_count = {}
    for game in schedule:
        home = game['home_name']
        away = game['away_name']
        if home in NAME_TO_FG and away in NAME_TO_FG:
            base_key = f"{NAME_TO_FG[away]}_{NAME_TO_FG[home]}"
            pair_count[base_key] = pair_count.get(base_key, 0) + 1
            key = base_key if pair_count[base_key] == 1 else f"{base_key}_G2"
            statuses[key] = {
                'status': game['status'],
                'home_score': game.get('home_score', '-'),
                'away_score': game.get('away_score', '-')
            }
    return statuses

def get_mlb_odds():
    url = f'https://api.the-odds-api.com/v4/sports/baseball_mlb/odds/?apiKey={API_KEY}&regions=us&markets=h2h,totals&oddsFormat=american&bookmakers=draftkings,fanduel'
    response = requests.get(url)
    data = response.json()
    return data

def get_park_factor(home_team, season=2025):
    pf_df = PARK_FACTORS[PARK_FACTORS['Season'] == season]
    team_name = next((k for k, v in PARK_FACTOR_TEAM_MAP.items() if v == home_team), None)
    if team_name is None:
        return 1.0
    row = pf_df[pf_df['Team'] == team_name]
    if row.empty:
        return 1.0
    return row['1yr'].values[0] / 100

# Load offense data
offense_2021 = pd.read_csv('season_stats/offense_2021.csv')
offense_2022 = pd.read_csv('season_stats/offense_2022.csv')
offense_2023 = pd.read_csv('season_stats/offense_2023.csv')
offense_2024 = pd.read_csv('season_stats/offense_2024.csv')
offense_2025 = pd.read_csv('season_stats/offense_2025.csv')
offense_2026 = pd.read_csv('season_stats/offense_2026.csv')

# Assign years
offense_2021['year'] = 2021
offense_2022['year'] = 2022
offense_2023['year'] = 2023
offense_2024['year'] = 2024
offense_2025['year'] = 2025
offense_2026['year'] = 2026

# Combine offense
offense = pd.concat([offense_2021, offense_2022, offense_2023, offense_2024,offense_2025,offense_2026])

# Load pitching data
pitching_2021 = pd.read_csv('season_stats/pitching_2021.csv')
pitching_2022 = pd.read_csv('season_stats/pitching_2022.csv')
pitching_2023 = pd.read_csv('season_stats/pitching_2023.csv')
pitching_2024 = pd.read_csv('season_stats/pitching_2024.csv')
pitching_2025 = pd.read_csv('season_stats/pitching_2025.csv')
pitching_2026 = pd.read_csv('season_stats/pitching_2026.csv')

# Assign years
pitching_2021['year'] = 2021
pitching_2022['year'] = 2022
pitching_2023['year'] = 2023
pitching_2024['year'] = 2024
pitching_2025['year'] = 2025
pitching_2026['year'] = 2026

# Combine pitching
pitching = pd.concat([pitching_2021, pitching_2022, pitching_2023, pitching_2024,pitching_2025,pitching_2026])

# Normalize column names
offense.columns = offense.columns.str.strip().str.lower().str.replace('+', 'plus', regex=False).str.replace('/', '_', regex=False).str.replace(' ', '_', regex=False).str.replace('%', '_pct', regex=False).str.replace('(', '', regex=False).str.replace(')', '', regex=False)
pitching.columns = pitching.columns.str.strip().str.lower().str.replace('+', 'plus', regex=False).str.replace('/', '_', regex=False).str.replace(' ', '_', regex=False).str.replace('%', '_pct', regex=False).str.replace('(', '', regex=False).str.replace(')', '', regex=False)

# Calculate runs per game for each team
offense['r_per_game'] = offense['r'] / offense['tg']
offense['sample_weight'] = (offense['tg']/162).clip(0,1)

# Calculate league average runs per game for each year
league_avg = offense.groupby('year')['r_per_game'].mean().reset_index()
league_avg.columns = ['year', 'lg_avg_runs']

# Merge league average into offense data
offense = offense.merge(league_avg, on='year')

# Calculate offensive rating relative to league average
offense['off_rating'] = offense['r_per_game']/offense['lg_avg_runs']

# Calculate league average FIP per year
lg_fip = pitching.groupby('year')['fip'].mean().reset_index()
lg_fip.columns = ['year','lg_avg_fip']

# Merge into pitching data
pitching = pitching.merge(lg_fip,on='year')

# Calculate pitching rating - not we invert it so better pitching = higher number
# A team with FIP below league average is BETTER, so we flip the ratio
pitching['pitch_rating'] = pitching['lg_avg_fip'] / pitching['fip']

model_data = offense[['team','year','r_per_game','off_rating']].merge(pitching[['team','year','pitch_rating']],on=['team','year'])

# Calculate lamba = expected runs scored against average pitching
# Adjusted for the opponent pitching quality
model_data = model_data.merge(league_avg, on='year')
model_data['lambda'] = model_data['off_rating']*model_data['lg_avg_runs']
  
# Calculate expected runs for both teams - New Beginning
def calculate_matchup(home_team, away_team, year, rolling=None, starters=None, injury_adj=None, platoon_data=None):
    if rolling and home_team in rolling and away_team in rolling:
        lg_avg_runs = league_avg[league_avg['year'] == 2026]['lg_avg_runs'].values[0]

        home_off_rating, home_pit_rating = calculate_team_ratings(
            home_team, rolling, all_woba, all_fip, team_xwoba, lg_avgs, lg_avg_runs, injury_adj)
        away_off_rating, away_pit_rating = calculate_team_ratings(
            away_team, rolling, all_woba, all_fip, team_xwoba, lg_avgs, lg_avg_runs, injury_adj)
        
        # Get park factor for home stadium
        park_factor = get_park_factor(home_team)
        
        # 7-day weighted 70%, 15-day 30% (backtest optimized)
        home_pitch_roll = rolling[home_team]['pitch_7'] * 0.70 + rolling[home_team]['pitch_15'] * 0.30
        away_pitch_roll = rolling[away_team]['pitch_7'] * 0.70 + rolling[away_team]['pitch_15'] * 0.30

        if starters and home_team in starters and starters[home_team]:
            home_innings = starters[home_team]['innings']
            home_ra9 = starters[home_team]['ra9']
            if home_innings >= MIN_INNINGS:
                home_starter_ra9 = min(home_ra9, MAX_RA9)
                home_blended_pitch = (home_starter_ra9 * STARTER_WEIGHT) + (home_pitch_roll * BULLPEN_WEIGHT)
            else:
                home_blended_pitch = home_pitch_roll
        else:
            home_blended_pitch = home_pitch_roll

        if starters and away_team in starters and starters[away_team]:
            away_innings = starters[away_team]['innings']
            away_ra9 = starters[away_team]['ra9']
            if away_innings >= MIN_INNINGS:
                away_starter_ra9 = min(away_ra9, MAX_RA9)
                away_blended_pitch = (away_starter_ra9 * STARTER_WEIGHT) + (away_pitch_roll * BULLPEN_WEIGHT)
            else:
                away_blended_pitch = away_pitch_roll
        else:
            away_blended_pitch = away_pitch_roll

        # Convert blended pitch to rating
        home_starter_pit_rating = lg_avg_runs / home_blended_pitch if home_blended_pitch > 0 else 1.0
        away_starter_pit_rating = lg_avg_runs / away_blended_pitch if away_blended_pitch > 0 else 1.0

        # Blend starter rating with full pitching rating
        home_final_pit = (home_starter_pit_rating * 0.50) + (home_pit_rating * 0.50)
        away_final_pit = (away_starter_pit_rating * 0.50) + (away_pit_rating * 0.50)

        # Calculate lambdas with park factor
        home_lambda = (home_off_rating / away_final_pit) * lg_avg_runs * park_factor
        away_lambda = (away_off_rating / home_final_pit) * lg_avg_runs * park_factor

        # Apply platoon factors when confirmed lineup is available
        if platoon_data:
            home_lambda *= platoon_data[0]
            away_lambda *= platoon_data[1]

    else:
        home_off = model_data[(model_data['team'] == home_team) & (model_data['year'] == year)]['lambda'].values[0]
        away_off = model_data[(model_data['team'] == away_team) & (model_data['year'] == year)]['lambda'].values[0]
        home_pitch = model_data[(model_data['team'] == home_team) & (model_data['year'] == year)]['pitch_rating'].values[0]
        away_pitch = model_data[(model_data['team'] == away_team) & (model_data['year'] == year)]['pitch_rating'].values[0]
        home_lambda = home_off / away_pitch
        away_lambda = away_off / home_pitch


    home_scores = np.random.poisson(home_lambda, NUM_SIMULATIONS)
    away_scores = np.random.poisson(away_lambda, NUM_SIMULATIONS)

    ties = home_scores == away_scores
    extra_home = np.random.poisson(home_lambda * 0.111, ties.sum())
    extra_away = np.random.poisson(away_lambda * 0.111, ties.sum())

    home_scores[ties] += extra_home
    away_scores[ties] += extra_away

    home_wins_sim = np.sum(home_scores > away_scores)
    away_wins_sim = np.sum(away_scores > home_scores)
    total_runs = home_scores + away_scores

    remaining = NUM_SIMULATIONS - home_wins_sim - away_wins_sim
    home_win_pct = (home_wins_sim + remaining / 2) / NUM_SIMULATIONS
    away_win_pct = (away_wins_sim + remaining / 2) / NUM_SIMULATIONS

    avg_total = np.mean(total_runs)

    return home_win_pct, away_win_pct, avg_total, home_lambda, away_lambda

def calculate_team_woba(team_id, season=2026):
    try:
        stats = statsapi.get('team_stats', {
            'teamId': team_id,
            'stats': 'season',
            'group': 'hitting',
            'season': season
        })
        s = stats['stats'][0]['splits'][0]['stat']
        
        bb = float(s.get('baseOnBalls', 0))
        hbp = float(s.get('hitByPitch', 0))
        hits = float(s.get('hits', 0))
        doubles = float(s.get('doubles', 0))
        triples = float(s.get('triples', 0))
        hr = float(s.get('homeRuns', 0))
        ab = float(s.get('atBats', 0))
        ibb = float(s.get('intentionalWalks', 0))
        sf = float(s.get('sacFlies', 0))
        pa = float(s.get('plateAppearances', 0))
        
        singles = hits - doubles - triples - hr
        k = float(s.get('strikeOuts', 0))

        numerator = (WOBA_WEIGHTS['wBB'] * bb +
                    WOBA_WEIGHTS['wHBP'] * hbp +
                    WOBA_WEIGHTS['w1B'] * singles +
                    WOBA_WEIGHTS['w2B'] * doubles +
                    WOBA_WEIGHTS['w3B'] * triples +
                    WOBA_WEIGHTS['wHR'] * hr)

        denominator = ab + bb - ibb + sf + hbp

        if denominator == 0:
            return None

        woba = numerator / denominator
        k_pct = k / pa if pa > 0 else 0
        bb_pct = bb / pa if pa > 0 else 0
        babip_denom = ab - k - hr + sf
        babip = (hits - hr) / babip_denom if babip_denom > 0 else None

        return {
            'woba':  round(woba, 3),
            'k_pct': round(k_pct, 3),
            'bb_pct': round(bb_pct, 3),
            'babip': round(babip, 3) if babip is not None else None,
            'pa':    int(pa),
        }
    except Exception as e:
        return None

def compare_to_market(game, model_home_prob, model_away_prob):
    home_team = game['home_team']
    away_team = game['away_team']
    
    if len(game['bookmakers']) == 0:
        return
    
    bookmaker = game['bookmakers'][0]
    outcomes = bookmaker['markets'][0]['outcomes']
    
    for outcome in outcomes:
        if outcome['name'] == home_team:
            market_home_line = outcome['price']
        else:
            market_away_line = outcome['price']
    
    market_home_prob = american_to_prob(market_home_line)
    market_away_prob = american_to_prob(market_away_line)
    
    home_edge = model_home_prob - market_home_prob
    away_edge = model_away_prob - market_away_prob
        
def get_team_logs(team_id, season=2026):
    try:
        stats=statsapi.get('team_stats', {'teamId':team_id,'stats':'gamelog','group':'hitting','season':season})
        games=stats['stats'][0]['splits']
        log = []
        for game in games:
            log.append({'date':game['date'],'runs':game['stat']['runs'],'atBats':game['stat']['atBats']})
        return log
    except Exception as e:
        print(f"Error fetching game log: {e}")
        return[]

def get_rolling_average(game_log, days=7, column='runs'):
    if not game_log or len(game_log) ==0:
        return None
    df = pd.DataFrame(game_log)
    df = df.sort_values('date')
    recent = df.tail(days)
    return recent[column].mean()
    
def get_pitching_game_logs(team_id,season=2026):
    try:
        stats=statsapi.get('team_stats',{'teamId': team_id,'stats':'gamelog','group':'pitching','season':season})
        games=stats['stats'][0]['splits']
        log = []
        for game in games:
            log.append({'date':game['date'],'runs_allowed':game['stat']['runs'],'inningsPitched': game['stat']['inningsPitched']})
        return log
    except Exception as e:
        print(f"Error fetching pitching log: {e}")
        return []


def get_all_team_ids():
    teams = statsapi.get('teams',{'sportId':1,'season':2026})
    teams_ids={}
    for team in teams['teams']:
        abbrev=team.get('abbreviation')
        if abbrev in TEAM_MAP:
            fg_abbrev=TEAM_MAP[abbrev]
            teams_ids[fg_abbrev]=team['id']
    return teams_ids

def get_all_team_woba(team_ids, season=2026):
    results = {}
    for fg_abbrev, team_id in team_ids.items():
        stats = calculate_team_woba(team_id, season)
        if stats:
            results[fg_abbrev] = stats
    return results

def get_team_xwoba(team_ids, season=2026):
    try:
        url = f"https://baseballsavant.mlb.com/leaderboard/expected_statistics?type=batter&year={season}&position=&team=&min=10&csv=true"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:138.0) Gecko/20100101 Firefox/138.0'}
        response = requests.get(url, headers=headers)
        xwoba_df = pd.read_csv(StringIO(response.text))
        
        player_team = {}
        for fg_abbrev, team_id in team_ids.items():
            try:
                roster_data = statsapi.get('team_roster', {
                    'teamId': team_id,
                    'rosterType': 'active'
                })
                for player in roster_data['roster']:
                    player_team[player['person']['id']] = fg_abbrev
            except:
                continue
        
        xwoba_df['team'] = xwoba_df['player_id'].map(player_team)
        xwoba_df = xwoba_df.dropna(subset=['team'])
        
        team_xwoba = xwoba_df.groupby('team').apply(
            lambda x: pd.Series({
                'xwoba': (x['est_woba'] * x['pa']).sum() / x['pa'].sum(),
                'woba_savant': (x['woba'] * x['pa']).sum() / x['pa'].sum()
            })
        ).reset_index()
        
        return dict(zip(team_xwoba['team'], team_xwoba['xwoba']))
    except Exception as e:
        print(f"xwOBA fetch failed: {e}")
        return {}

def calculate_lg_hr_fb(team_ids, season=2026):
    total_hr = 0
    total_fb = 0
    for fg_abbrev, team_id in team_ids.items():
        try:
            stats = statsapi.get('team_stats', {
                'teamId': team_id,
                'stats': 'season',
                'group': 'pitching',
                'season': season
            })
            s = stats['stats'][0]['splits'][0]['stat']
            hr = float(s.get('homeRuns', 0))
            air_outs = float(s.get('airOuts', 0))
            total_hr += hr
            total_fb += (air_outs + hr)
        except:
            continue
    if total_fb == 0:
        return 0.115
    return total_hr / total_fb

def calculate_team_fip(team_id, lg_hr_fb, season=2026):
    try:
        stats = statsapi.get('team_stats', {
            'teamId': team_id,
            'stats': 'season',
            'group': 'pitching',
            'season': season
        })
        s = stats['stats'][0]['splits'][0]['stat']
        
        hr = float(s.get('homeRuns', 0))
        bb = float(s.get('baseOnBalls', 0))
        hbp = float(s.get('hitBatsmen', 0))
        k = float(s.get('strikeOuts', 0))
        air_outs = float(s.get('airOuts', 0))
        bf = float(s.get('battersFaced', 1))
        ip_str = s.get('inningsPitched', '0')
        ip = ip_to_float(ip_str)

        if ip == 0:
            return None
        
        fb = air_outs + hr
        expected_hr = fb * lg_hr_fb
        
        fip = ((13 * hr) + (3 * (bb + hbp)) - (2 * k)) / ip + FIP_CONSTANT
        xfip = ((13 * expected_hr) + (3 * (bb + hbp)) - (2 * k)) / ip + FIP_CONSTANT
        
        k_pct = k / bf
        bb_pct = bb / bf
        
        return {
            'fip':      round(fip, 3),
            'xfip':     round(xfip, 3),
            'k_pct':    round(k_pct, 3),
            'bb_pct':   round(bb_pct, 3),
            'lg_hr_fb': round(lg_hr_fb, 4),
            'ip':       round(ip, 1),
        }
    except Exception as e:
        return None

def get_all_team_fip(team_ids, season=2026):
    lg_hr_fb = calculate_lg_hr_fb(team_ids, season)
    results = {}
    for fg_abbrev, team_id in team_ids.items():
        stats = calculate_team_fip(team_id, lg_hr_fb, season)
        if stats:
            results[fg_abbrev] = stats
    return results

def get_injury_adjustments(team_ids, all_woba, all_fip, lg_avgs, season=2026):
    """
    Returns per-team rating adjustments for players placed on IL within the last 14 days.
    Adjustment tapers linearly from full (day 1) to zero (day 14) as rolling averages catch up.
    Format: {team_fg: {'off_adj': float, 'pit_adj': float, 'notes': [str]}}
    """
    today = date.today()
    season_start = f'{season}-03-01'

    try:
        url  = (f"https://statsapi.mlb.com/api/v1/transactions"
                f"?startDate={season_start}&endDate={today.strftime('%Y-%m-%d')}&sportId=1")
        resp = requests.get(url, timeout=10).json()
    except Exception as e:
        print(f"Injury transactions fetch failed: {e}")
        return {}

    transactions = resp.get('transactions', [])

    # Build placed/activated sets by scanning in date order
    il_placed    = {}   # player_id -> {name, team_fg, date}
    il_activated = set()

    for t in sorted(transactions, key=lambda x: x.get('date', '')):
        desc   = t.get('description', '')
        person = t.get('person', {})
        pid    = person.get('id')
        if not pid:
            continue

        # Resolve team abbreviation — try multiple fields
        mlb_abbrev = (
            (t.get('toTeam')   or {}).get('abbreviation') or
            (t.get('fromTeam') or {}).get('abbreviation') or
            (t.get('team')     or {}).get('abbreviation')
        )
        team_fg = TEAM_MAP.get(mlb_abbrev) if mlb_abbrev else None

        if 'Placed' in desc and 'Injured List' in desc:
            il_placed[pid] = {
                'name':    person.get('fullName', 'Unknown'),
                'team_fg': team_fg,
                'date':    t.get('date', ''),
            }
        elif 'Activated' in desc and 'Injured List' in desc:
            il_activated.add(pid)

    # Currently on IL = placed but not yet activated
    current_il = {pid: info for pid, info in il_placed.items()
                  if pid not in il_activated}

    adjustments = {}

    for pid, info in current_il.items():
        team_fg = info['team_fg']
        if not team_fg or team_fg not in team_ids:
            continue

        try:
            il_date   = datetime.strptime(info['date'][:10], '%Y-%m-%d').date()
        except (ValueError, AttributeError):
            continue

        days_on_il = (today - il_date).days
        if days_on_il >= 14:
            continue  # rolling averages have fully adjusted

        taper = (14 - days_on_il) / 14.0  # 1.0 on day 1 → 0.0 on day 14

        if team_fg not in adjustments:
            adjustments[team_fg] = {'off_adj': 0.0, 'pit_adj': 0.0, 'notes': []}

        # --- Hitting impact ---
        try:
            hdata = statsapi.player_stat_data(pid, group='hitting', type='season', sportId=1)
            if hdata and hdata.get('stats'):
                s  = hdata['stats'][0]['stats']
                pa = float(s.get('plateAppearances', 0))
                if pa >= 30:
                    hits    = float(s.get('hits', 0))
                    doubles = float(s.get('doubles', 0))
                    triples = float(s.get('triples', 0))
                    hr      = float(s.get('homeRuns', 0))
                    bb      = float(s.get('baseOnBalls', 0))
                    hbp     = float(s.get('hitByPitch', 0))
                    ibb     = float(s.get('intentionalWalks', 0))
                    sf      = float(s.get('sacFlies', 0))
                    ab      = float(s.get('atBats', 0))
                    singles = hits - doubles - triples - hr
                    woba_d  = ab + bb - ibb + sf + hbp
                    if woba_d > 0:
                        player_woba = (
                            WOBA_WEIGHTS['wBB']  * (bb - ibb) +
                            WOBA_WEIGHTS['wHBP'] * hbp +
                            WOBA_WEIGHTS['w1B']  * singles +
                            WOBA_WEIGHTS['w2B']  * doubles +
                            WOBA_WEIGHTS['w3B']  * triples +
                            WOBA_WEIGHTS['wHR']  * hr
                        ) / woba_d
                        lg_woba  = lg_avgs['lg_woba']
                        team_pa  = all_woba.get(team_fg, {}).get('pa', 0)
                        if lg_woba > 0 and team_pa > 0:
                            pa_share  = pa / team_pa  # no cap — PA share is naturally bounded (~10-15% for elite regulars)
                            woba_rate = player_woba / lg_woba
                            off_adj   = (woba_rate - 1.0) * pa_share * taper
                            adjustments[team_fg]['off_adj'] += off_adj
                            adjustments[team_fg]['notes'].append(
                                f"{info['name']} OUT {days_on_il}d — off {off_adj:+.3f}"
                            )
        except Exception:
            pass

        # --- Pitching impact ---
        try:
            pdata = statsapi.player_stat_data(pid, group='pitching', type='season', sportId=1)
            if pdata and pdata.get('stats'):
                s  = pdata['stats'][0]['stats']
                ip = ip_to_float(s.get('inningsPitched', '0'))
                if ip >= 10:
                    hr  = float(s.get('homeRuns', 0))
                    bb  = float(s.get('baseOnBalls', 0))
                    hbp = float(s.get('hitBatsmen', 0))
                    k   = float(s.get('strikeOuts', 0))
                    player_fip = ((13 * hr) + (3 * (bb + hbp)) - (2 * k)) / ip + FIP_CONSTANT
                    lg_fip   = lg_avgs['lg_fip']
                    team_ip  = all_fip.get(team_fg, {}).get('ip', 0)
                    if player_fip > 0 and team_ip > 0:
                        ip_share  = min(ip / team_ip, 0.25)
                        fip_rate  = lg_fip / player_fip
                        pit_adj   = (fip_rate - 1.0) * ip_share * taper
                        adjustments[team_fg]['pit_adj'] += pit_adj
                        adjustments[team_fg]['notes'].append(
                            f"{info['name']} OUT {days_on_il}d — pit {pit_adj:+.3f}"
                        )
        except Exception:
            pass

    return adjustments

MLB_PEOPLE_URL = 'https://statsapi.mlb.com/api/v1/people'

def get_confirmed_lineups(game_date):
    """
    Returns confirmed lineups and pitcher IDs for today's games via MLB API.
    Format: {(home_fg, away_fg): {home_lineup, away_lineup, home_pitcher_id, away_pitcher_id}}
    Returns {} if no lineups confirmed yet.
    """
    try:
        resp = requests.get(
            'https://statsapi.mlb.com/api/v1/schedule',
            params={'date': game_date, 'sportId': 1, 'hydrate': 'lineups,probablePitcher'},
            timeout=10
        ).json()
        result = {}
        for date_entry in resp.get('dates', []):
            for game in date_entry.get('games', []):
                teams     = game.get('teams', {})
                home_name = teams.get('home', {}).get('team', {}).get('name', '')
                away_name = teams.get('away', {}).get('team', {}).get('name', '')
                if home_name not in NAME_TO_FG or away_name not in NAME_TO_FG:
                    continue
                home_fg = NAME_TO_FG[home_name]
                away_fg = NAME_TO_FG[away_name]
                lineups     = game.get('lineups', {})
                home_lineup = [p['id'] for p in lineups.get('homePlayers', [])]
                away_lineup = [p['id'] for p in lineups.get('awayPlayers', [])]
                # Read probable pitchers before the skip check so pre-lineup
                # games with a known starter still enter the lineups dict.
                home_pid = teams.get('home', {}).get('probablePitcher', {}).get('id')
                away_pid = teams.get('away', {}).get('probablePitcher', {}).get('id')
                # Skip only if there is truly nothing useful (no lineup AND no probables)
                if not home_lineup and not away_lineup and not home_pid and not away_pid:
                    continue
                result[(home_fg, away_fg)] = {
                    'home_lineup':    home_lineup,
                    'away_lineup':    away_lineup,
                    'home_pitcher_id': home_pid,
                    'away_pitcher_id': away_pid,
                }
        return result
    except Exception as e:
        print(f"Lineup fetch failed: {e}")
        return {}

def get_players_hand(player_ids):
    """Batch fetch bat/pitch handedness for a list of player IDs in one API call."""
    if not player_ids:
        return {}
    try:
        ids_str = ','.join(str(p) for p in player_ids)
        resp = requests.get(MLB_PEOPLE_URL, params={'personIds': ids_str}, timeout=10).json()
        result = {}
        for person in resp.get('people', []):
            result[person['id']] = {
                'bat_hand':   person.get('batSide',   {}).get('code', 'R'),
                'pitch_hand': person.get('pitchHand', {}).get('code', 'R'),
            }
        return result
    except Exception:
        return {}

def get_player_split_woba(player_id, season, min_pa=30):
    """Returns {vs_lhp: woba, vs_rhp: woba} for a batter. None if insufficient data."""
    try:
        resp = requests.get(
            f'{MLB_PEOPLE_URL}/{player_id}/stats',
            params={'stats': 'statSplits', 'group': 'hitting', 'season': season, 'sportId': 1},
            timeout=8
        ).json()
        splits = resp.get('stats', [{}])[0].get('splits', [])
        result = {}
        for split in splits:
            code = split.get('split', {}).get('code', '')
            if code not in ('vl', 'vr'):
                continue
            s  = split.get('stat', {})
            pa = int(s.get('plateAppearances', 0))
            if pa < min_pa:
                continue
            hits    = float(s.get('hits', 0))
            doubles = float(s.get('doubles', 0))
            triples = float(s.get('triples', 0))
            hr      = float(s.get('homeRuns', 0))
            bb      = float(s.get('baseOnBalls', 0))
            hbp     = float(s.get('hitByPitch', 0))
            ibb     = float(s.get('intentionalWalks', 0))
            sf      = float(s.get('sacFlies', 0))
            ab      = float(s.get('atBats', 0))
            singles = hits - doubles - triples - hr
            denom   = ab + bb - ibb + sf + hbp
            if denom > 0:
                woba = (
                    WOBA_WEIGHTS['wBB']  * (bb - ibb) +
                    WOBA_WEIGHTS['wHBP'] * hbp        +
                    WOBA_WEIGHTS['w1B']  * singles     +
                    WOBA_WEIGHTS['w2B']  * doubles     +
                    WOBA_WEIGHTS['w3B']  * triples     +
                    WOBA_WEIGHTS['wHR']  * hr
                ) / denom
                result['vs_lhp' if code == 'vl' else 'vs_rhp'] = round(woba, 3)
        return result or None
    except Exception:
        return None

def get_pitcher_split_fip(pitcher_id, season, min_ip=5):
    """Returns {vs_lhb: fip, vs_rhb: fip} for a pitcher. None if insufficient data."""
    try:
        resp = requests.get(
            f'{MLB_PEOPLE_URL}/{pitcher_id}/stats',
            params={'stats': 'statSplits', 'group': 'pitching', 'season': season, 'sportId': 1},
            timeout=8
        ).json()
        splits = resp.get('stats', [{}])[0].get('splits', [])
        result = {}
        for split in splits:
            code = split.get('split', {}).get('code', '')
            if code not in ('vl', 'vr'):
                continue
            s  = split.get('stat', {})
            ip = ip_to_float(s.get('inningsPitched', '0'))
            if ip < min_ip:
                continue
            hr  = float(s.get('homeRuns', 0))
            bb  = float(s.get('baseOnBalls', 0))
            hbp = float(s.get('hitBatsmen', 0))
            k   = float(s.get('strikeOuts', 0))
            fip = ((13 * hr) + (3 * (bb + hbp)) - (2 * k)) / ip + FIP_CONSTANT
            result['vs_lhb' if code == 'vl' else 'vs_rhb'] = round(fip, 3)
        return result or None
    except Exception:
        return None

def get_all_platoon_splits(lineup_data, season=2026):
    """
    Fetches and caches platoon splits for all players in confirmed lineups.
    Returns {bat_{id}: {splits, hand}, pit_{id}: {splits, hand}} cache dict.
    Cache TTL: 6 hours (splits don't change intraday).
    """
    cache_file = 'cache/splits_cache.json'
    cache = {}
    if os.path.exists(cache_file):
        age_hours = (datetime.now().timestamp() - os.path.getmtime(cache_file)) / 3600
        if age_hours < 6:
            with open(cache_file, 'r') as f:
                cache = json.load(f)

    all_batter_ids  = set()
    all_pitcher_ids = set()
    for gd in lineup_data.values():
        all_batter_ids.update(gd.get('home_lineup', []))
        all_batter_ids.update(gd.get('away_lineup', []))
        if gd.get('home_pitcher_id'): all_pitcher_ids.add(gd['home_pitcher_id'])
        if gd.get('away_pitcher_id'): all_pitcher_ids.add(gd['away_pitcher_id'])

    # Batch-fetch handedness for anyone not cached
    missing_hand = [p for p in (all_batter_ids | all_pitcher_ids)
                    if f"bat_{p}" not in cache and f"pit_{p}" not in cache]
    hand_info = get_players_hand(missing_hand)

    # Fetch batter splits
    for pid in all_batter_ids:
        key = f"bat_{pid}"
        if key not in cache:
            splits = get_player_split_woba(pid, season)
            hand   = hand_info.get(pid, {}).get('bat_hand', 'R')
            cache[key] = {'splits': splits, 'hand': hand}

    # Fetch pitcher splits + handedness
    for pid in all_pitcher_ids:
        key = f"pit_{pid}"
        if key not in cache:
            splits = get_pitcher_split_fip(pid, season)
            hand   = hand_info.get(pid, {}).get('pitch_hand', 'R')
            cache[key] = {'splits': splits, 'hand': hand}

    with open(cache_file, 'w') as f:
        json.dump(cache, f)

    return cache

def compute_platoon_factors(home_lineup, away_lineup, home_pitcher_id, away_pitcher_id,
                             splits_cache, lg_avgs):
    """
    Returns (home_off_factor, away_off_factor, matchup_notes).
    Factor > 1.0 = favorable platoon matchup, < 1.0 = unfavorable.
    No artificial cap — PA/IP minimums already guard against small samples.
    """
    lg_woba = lg_avgs.get('lg_woba', 0.315)

    def lineup_factor(lineup_ids, pitcher_id, cache):
        pit_info     = cache.get(f"pit_{pitcher_id}", {})
        pitcher_hand = pit_info.get('hand', 'R')
        split_key    = 'vs_rhp' if pitcher_hand == 'R' else 'vs_lhp'
        woba_vals = []
        for pid in lineup_ids:
            bat_info = cache.get(f"bat_{pid}", {})
            splits   = bat_info.get('splits') or {}
            woba     = splits.get(split_key)
            if woba and woba > 0:
                woba_vals.append(woba)
        if len(woba_vals) < 3:   # need ≥3 batters with qualifying data
            return 1.0, pitcher_hand, len(woba_vals)
        factor = sum(woba_vals) / len(woba_vals) / lg_woba
        return factor, pitcher_hand, len(woba_vals)

    home_factor, away_hand, home_n = lineup_factor(home_lineup, away_pitcher_id, splits_cache)
    away_factor, home_hand, away_n = lineup_factor(away_lineup, home_pitcher_id, splits_cache)

    notes = (
        f"vs {'L' if away_hand == 'L' else 'R'}HP ({home_n} batters) | "
        f"vs {'L' if home_hand == 'L' else 'R'}HP ({away_n} batters)"
    )
    return home_factor, away_factor, notes

def get_all_rolling_averages(team_ids, days_short=7, days_long=15):
    results = {}
    for fg_abbrev, team_id in team_ids.items():
        hitting_log = get_team_logs(team_id)
        pitching_log = get_pitching_game_logs(team_id)
        hit_short = get_rolling_average(hitting_log, days=days_short)
        hit_long = get_rolling_average(hitting_log, days=days_long)
        pitch_short = get_rolling_average(pitching_log, days=days_short, column='runs_allowed')
        pitch_long = get_rolling_average(pitching_log, days=days_long, column='runs_allowed')
        results[fg_abbrev] = {
            'hit_7': hit_short,
            'hit_15': hit_long,
            'pitch_7': pitch_short,
            'pitch_15': pitch_long
        }
    return results

from datetime import datetime
run_time = datetime.now().strftime('%H:%M')
todays_predictions = []

def get_yesterdays_results():
    yesterday = (date.today() - pd.Timedelta(days=1)).strftime('%Y-%m-%d')
    schedule = statsapi.schedule(date=yesterday)
    schedule = sorted(schedule, key=lambda g: g.get('game_datetime', ''))
    results = {}
    postponed = set()
    pair_count = {}
    for game in schedule:
        home = game['home_name']
        away = game['away_name']
        if home not in NAME_TO_FG or away not in NAME_TO_FG:
            continue
        base_key = f"{NAME_TO_FG[away]}_{NAME_TO_FG[home]}"
        pair_count[base_key] = pair_count.get(base_key, 0) + 1
        key = base_key if pair_count[base_key] == 1 else f"{base_key}_G2"
        if game['status'] == 'Final':
            home_score = game['home_score']
            away_score = game['away_score']
            results[key] = {
                'home': NAME_TO_FG[home],
                'away': NAME_TO_FG[away],
                'home_score': home_score,
                'away_score': away_score,
                'winner': NAME_TO_FG[home] if home_score > away_score else NAME_TO_FG[away]
            }
        elif game['status'] in ('Postponed', 'Cancelled'):
            postponed.add(key)
    return results, postponed

yesterdays_results, yesterdays_postponed = get_yesterdays_results()

_NO_BET_VALS = {'No Bet', 'No Edge', 'No Line', 'nan', 'None', '', 'N/A', 'CANCELLED'}

def _is_active_bet(val):
    return pd.notna(val) and str(val) not in _NO_BET_VALS

def compute_run_change(prev_bet, new_bet):
    prev = _is_active_bet(prev_bet)
    new  = _is_active_bet(new_bet)
    if not prev and not new:   return '—'
    if not prev and new:       return 'EDGE GAINED'
    if prev and not new:       return 'EDGE LOST'
    if str(prev_bet) != str(new_bet): return 'PICK CHANGED'
    return '—'

def get_final_run(game_time_str, row):
    """Return the first run that had an active bet for this game.
    Falls back to the last scheduled run before game time if no active bets found yet.
    Schedule (MST): run1=6PM, run2=11PM, run3=5AM, run4=11AM, run5=3PM.
    """
    for run in range(1, 6):
        if _is_active_bet(row.get(f'run{run}_bet_team')):
            return run
    # No active bets yet — return last scheduled run before game time as placeholder
    try:
        hour = int(str(game_time_str).split(':')[0])
    except Exception:
        hour = 19
    if hour < 11:  return 3
    if hour < 15:  return 4
    return 5

def log_predictions(predictions, results, postponed=None, statuses=None, active_game_keys=None,
                    run_num=None, run_window=None):
    log_file = 'predictions/predictions_log.csv'

    def _game_num_str(val):
        return '1' if (val is None or pd.isna(val) or str(val) in ('', 'nan', 'None')) else str(val)

    if run_num is None or run_window is None:
        run_num, run_window, _ = _compute_run_context()
    today_str = date.today().strftime('%Y-%m-%d')
    yesterday_str = (date.today() - pd.Timedelta(days=1)).strftime('%Y-%m-%d')

    if not predictions:
        return

    columns = [
        'date', 'game_time', 'game_num', 'home_team', 'away_team', 'bet_type', 'model_strategy',
        'run_window_mst',
        'run1_bet_team', 'run1_model_pct', 'run1_book_line', 'run1_edge',
        'run2_bet_team', 'run2_model_pct', 'run2_book_line', 'run2_edge', 'run2_change',
        'run3_bet_team', 'run3_model_pct', 'run3_book_line', 'run3_edge', 'run3_change',
        'run4_bet_team', 'run4_model_pct', 'run4_book_line', 'run4_edge', 'run4_change',
        'run5_bet_team', 'run5_model_pct', 'run5_book_line', 'run5_edge', 'run5_change',
        'run6_bet_team', 'run6_model_pct', 'run6_book_line', 'run6_edge', 'run6_change',
        'run7_bet_team', 'run7_model_pct', 'run7_book_line', 'run7_edge', 'run7_change',
        'run8_bet_team', 'run8_model_pct', 'run8_book_line', 'run8_edge', 'run8_change',
        'run9_bet_team', 'run9_model_pct', 'run9_book_line', 'run9_edge', 'run9_change',
        'final_run', 'final_bet_team', 'final_model_pct', 'final_book_line', 'final_edge',
        'raw_model_pct', 'home_platoon_factor', 'away_platoon_factor', 'platoon_confirmed',
        'recommended_stake',
        'actual_home_score', 'actual_away_score', 'winner', 'result', 'ou_direction',
    ]

    if os.path.exists(log_file):
        df = pd.read_csv(log_file, dtype=str)
        df = df.rename(columns={'home_score': 'actual_home_score', 'away_score': 'actual_away_score'})
        for col in columns:
            if col not in df.columns:
                df[col] = None
        # Normalize date column — old rows used M/D/YYYY, new rows use YYYY-MM-DD
        # Mixed formats break string sorting (e.g. "5/22/2026" > "2026-05-23")
        def _norm_date(d):
            if pd.isna(d) or str(d) in ('', 'nan', 'None'):
                return d
            try:
                return pd.to_datetime(str(d)).strftime('%Y-%m-%d')
            except Exception:
                return d
        df['date'] = df['date'].apply(_norm_date)

        has_scores = (
            df['actual_home_score'].notna() &
            ~df['actual_home_score'].isin(['', 'nan', 'None']) &
            df['actual_away_score'].notna() &
            ~df['actual_away_score'].isin(['', 'nan', 'None'])
        ).fillna(False).astype(bool)

        # Retroactive fix: game_num NaN for rows created before column existed → default to 1
        gnum_missing = df['game_num'].isna() | df['game_num'].isin(['', 'nan', 'None'])
        df.loc[gnum_missing, 'game_num'] = '1'

        # Retroactive fix: ML No Bet rows with scores stuck at PENDING → N/A
        ml_nobet_mask = (
            (df['bet_type'] == 'Moneyline') &
            ~df['final_bet_team'].apply(_is_active_bet).astype(bool) &
            (df['result'].isin(['PENDING', '']) | df['result'].isna()) &
            has_scores
        )
        df.loc[ml_nobet_mask, 'result'] = 'N/A'

        # Retroactive fix: O/U rows with scores but result stuck at PENDING
        ou_fix_mask = (
            (df['bet_type'] == 'Over/Under') &
            (df['result'].isin(['PENDING', '']) | df['result'].isna()) &
            has_scores
        )
        for idx in df[ou_fix_mask].index:
            row = df.loc[idx]
            bet_team = row.get('final_bet_team')
            if not _is_active_bet(bet_team):
                for run in range(9, 0, -1):
                    t = row.get(f'run{run}_bet_team')
                    if _is_active_bet(t):
                        bet_team = t
                        break
            book_line = row.get('final_book_line')
            if not book_line or str(book_line) in ('', 'nan', 'None'):
                for run in range(9, 0, -1):
                    bl = row.get(f'run{run}_book_line')
                    if bl and str(bl) not in ('', 'nan', 'None'):
                        book_line = bl
                        break
            try:
                fake_result = {
                    'home_score': row['actual_home_score'],
                    'away_score': row['actual_away_score'],
                }
                pred = {'bet_type': 'Over/Under', 'bet_team': bet_team, 'book_line': book_line}
                new_result = evaluate_result(pred, fake_result)
                if new_result != 'PENDING':
                    df.at[idx, 'result'] = new_result
            except Exception:
                pass

        # Retroactive fix: compute ou_direction for all O/U rows that have scores but no direction
        if 'ou_direction' not in df.columns:
            df['ou_direction'] = None
        ou_dir_mask = (
            (df['bet_type'] == 'Over/Under') &
            (df['ou_direction'].isna() | df['ou_direction'].isin(['', 'nan', 'None']) |
             df['ou_direction'].str.contains('nan', na=False)) &
            has_scores
        )
        def _valid_pct(v):
            return pd.notna(v) and str(v) not in ('', 'nan', 'None')

        for idx in df[ou_dir_mask].index:
            row = df.loc[idx]
            model_total = next((row.get(f'run{r}_model_pct') for r in range(9, 0, -1)
                                if _valid_pct(row.get(f'run{r}_model_pct'))), None)
            if not _valid_pct(model_total):
                mt = row.get('final_model_pct')
                model_total = mt if _valid_pct(mt) else None
            direction = compute_ou_direction(model_total, row['actual_home_score'], row['actual_away_score'])
            if direction:
                df.at[idx, 'ou_direction'] = direction
    else:
        df = pd.DataFrame(columns=columns)

    # ── Pass 1: log today's predictions — always PENDING ──────────────────────
    for pred in predictions:
        game_date = pred['date']
        home_fg   = pred['home_fg']
        away_fg   = pred['away_fg']
        bet_type  = pred['bet_type']
        game_time = pred.get('game_time', '')
        game_num  = str(pred.get('game_num', 1) or 1)

        gn_col = df['game_num'].astype(str)
        if game_num == '1':
            game_num_mask = gn_col.isin(['1']) | df['game_num'].isna() | gn_col.isin(['None', ''])
        else:
            game_num_mask = gn_col == game_num

        mask = (
            (df['date']      == game_date) &
            (df['home_team'] == home_fg)   &
            (df['away_team'] == away_fg)   &
            (df['bet_type']  == bet_type)  &
            game_num_mask
        )

        if df[mask].empty:
            new_row = {col: None for col in columns}
            new_row['date']           = game_date
            new_row['game_time']      = game_time
            new_row['game_num']       = game_num
            new_row['home_team']      = home_fg
            new_row['away_team']      = away_fg
            new_row['bet_type']       = bet_type
            new_row['model_strategy'] = pred.get('model_strategy', 'SNIPER')
            new_row['run_window_mst']          = run_window
            new_row[f'run{run_num}_bet_team']  = pred['bet_team']
            new_row[f'run{run_num}_model_pct'] = pred['model_pct']
            new_row[f'run{run_num}_book_line'] = pred['book_line']
            new_row[f'run{run_num}_edge']      = pred['edge']
            new_row['raw_model_pct']           = pred.get('raw_model_pct', '')
            new_row['home_platoon_factor']     = pred.get('home_platoon_factor', '')
            new_row['away_platoon_factor']     = pred.get('away_platoon_factor', '')
            new_row['platoon_confirmed']       = pred.get('platoon_confirmed', 'No')
            new_row['recommended_stake']       = pred.get('recommended_stake', 0.0)
            final_run = get_final_run(game_time, new_row)
            new_row['final_run']       = str(final_run)
            new_row['final_bet_team']  = new_row.get(f'run{final_run}_bet_team')
            new_row['final_model_pct'] = new_row.get(f'run{final_run}_model_pct')
            new_row['final_book_line'] = new_row.get(f'run{final_run}_book_line')
            new_row['final_edge']      = new_row.get(f'run{final_run}_edge')
            new_row['result'] = 'PENDING'
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            idx = df[mask].index[0]
            # Backfill game_time and game_num if missing (rows created before columns existed)
            existing_gt = df.at[idx, 'game_time']
            if not existing_gt or str(existing_gt) in ('', 'nan', 'None'):
                df.at[idx, 'game_time'] = game_time
            existing_gn = df.at[idx, 'game_num']
            if pd.isna(existing_gn) or str(existing_gn) in ('', 'nan', 'None'):
                df.at[idx, 'game_num'] = game_num
            df.at[idx, 'run_window_mst'] = run_window
            if run_num > 1:
                prev_bet = df.at[idx, f'run{run_num - 1}_bet_team']
                df.at[idx, f'run{run_num}_change'] = compute_run_change(prev_bet, pred['bet_team'])
            df.at[idx, f'run{run_num}_bet_team']  = pred['bet_team']
            df.at[idx, f'run{run_num}_model_pct'] = pred['model_pct']
            df.at[idx, f'run{run_num}_book_line'] = pred['book_line']
            df.at[idx, f'run{run_num}_edge']      = pred['edge']
            # Update platoon fields — lineup may have been confirmed since last run
            df.at[idx, 'raw_model_pct']       = pred.get('raw_model_pct') or df.at[idx, 'raw_model_pct']
            df.at[idx, 'home_platoon_factor'] = pred.get('home_platoon_factor') or df.at[idx, 'home_platoon_factor']
            df.at[idx, 'away_platoon_factor'] = pred.get('away_platoon_factor') or df.at[idx, 'away_platoon_factor']
            if pred.get('platoon_confirmed') == 'Yes':
                df.at[idx, 'platoon_confirmed'] = 'Yes'
            if pred.get('recommended_stake') is not None:
                df.at[idx, 'recommended_stake'] = str(pred['recommended_stake'])
            gt = df.at[idx, 'game_time'] or game_time
            # Only update final_* if no active bet has been established yet —
            # final_* locks to the FIRST run with an active bet (the opening-line bet).
            if not _is_active_bet(df.at[idx, 'final_bet_team']):
                final_run = get_final_run(gt, df.loc[idx])
                df.at[idx, 'final_run']       = str(final_run)
                df.at[idx, 'final_bet_team']  = df.at[idx, f'run{final_run}_bet_team']
                df.at[idx, 'final_model_pct'] = df.at[idx, f'run{final_run}_model_pct']
                df.at[idx, 'final_book_line'] = df.at[idx, f'run{final_run}_book_line']
                df.at[idx, 'final_edge']      = df.at[idx, f'run{final_run}_edge']

    # ── Pass 2: fill actual results for yesterday's rows ──────────────────────
    # Completely separate from today's predictions — eliminates the bug where
    # the same two teams playing back-to-back days would inherit yesterday's result.
    for idx in df[df['date'] == yesterday_str].index:
        row      = df.loc[idx]
        base_key = f"{row['away_team']}_{row['home_team']}"
        game_num = _game_num_str(row.get('game_num'))
        game_key = base_key if game_num == '1' else f"{base_key}_G2"
        if game_key not in results:
            if postponed and game_key in postponed:
                df.at[idx, 'result'] = 'VOID'
            continue
        r = results[game_key]
        df.at[idx, 'actual_home_score'] = str(r['home_score'])
        df.at[idx, 'actual_away_score'] = str(r['away_score'])
        df.at[idx, 'winner']            = r['winner']
        final_pred = {
            'bet_type': row['bet_type'],
            'bet_team': row.get('final_bet_team'),
            'book_line': row.get('final_book_line'),
        }
        df.at[idx, 'result'] = evaluate_result(final_pred, r)

        if row['bet_type'] == 'Over/Under':
            model_total = next((row.get(f'run{rn}_model_pct') for rn in range(9, 0, -1)
                                if _valid_pct(row.get(f'run{rn}_model_pct'))), None)
            if not _valid_pct(model_total):
                mt = row.get('final_model_pct')
                model_total = mt if _valid_pct(mt) else None
            direction = compute_ou_direction(model_total, r['home_score'], r['away_score'])
            if direction:
                df.at[idx, 'ou_direction'] = direction

    # Safety net: today's rows are always PENDING with no scores, regardless of old data
    today_mask = df['date'] == today_str
    df.loc[today_mask, 'result']            = 'PENDING'
    df.loc[today_mask, 'actual_home_score'] = None
    df.loc[today_mask, 'actual_away_score'] = None
    df.loc[today_mask, 'winner']            = None
    df.loc[today_mask, 'ou_direction']      = None

    # Mark cancelled/postponed today rows — zero out bet info so they don't look active
    # Triggers if MLB API says Postponed/Cancelled OR game no longer appears in odds feed
    for idx in df[today_mask].index:
        row      = df.loc[idx]
        base_key = f"{row['away_team']}_{row['home_team']}"
        game_num = _game_num_str(row.get('game_num'))
        mlb_key  = base_key if game_num == '1' else f"{base_key}_G2"
        mlb_status    = (statuses or {}).get(mlb_key, {}).get('status', '')
        in_odds       = (active_game_keys is None) or (mlb_key in active_game_keys)
        live_or_done  = (mlb_status in ('Final', 'In Progress', 'Warmup', 'Pre-Game',
                                         'Game Over', 'Completed Early')
                         or mlb_status.startswith('Delayed')
                         or mlb_status.startswith('Suspended'))
        if not live_or_done and (mlb_status in ('Postponed', 'Cancelled') or not in_odds):
            df.at[idx, 'result'] = 'CANCELLED'
            for run in range(1, 10):
                df.at[idx, f'run{run}_bet_team'] = 'CANCELLED'
                df.at[idx, f'run{run}_edge']     = 'CANCELLED'
            df.at[idx, 'final_bet_team'] = 'CANCELLED'
            df.at[idx, 'final_edge']     = 'CANCELLED'

    df = df.reindex(columns=columns)
    df = df.sort_values(['date', 'game_time'], ascending=[False, True], na_position='last').reset_index(drop=True)
    df.to_csv(log_file, index=False)
    generate_excel_log(df)
    update_google_sheets(df)

def compute_ou_direction(model_total, actual_home, actual_away, close_threshold=1.0):
    """Returns HIGH/LOW/CLOSE with signed error — how accurate was the model's run total?"""
    try:
        projected = float(model_total)
        actual    = float(actual_home) + float(actual_away)
        if pd.isna(projected) or pd.isna(actual):
            return None
        error = projected - actual
        if abs(error) <= close_threshold:
            return f'CLOSE {error:+.1f}'
        elif error > 0:
            return f'HIGH {error:+.1f}'
        else:
            return f'LOW {error:+.1f}'
    except (ValueError, TypeError):
        return None

def evaluate_result(pred, game_result):
    bet_type  = pred.get('bet_type', '')
    bet_team  = pred.get('bet_team', '')
    bet_team  = str(bet_team) if (bet_team is not None and str(bet_team) != 'nan') else ''

    if bet_team in _NO_BET_VALS:
        return 'N/A'

    if bet_type == 'Moneyline':
        if game_result['winner'] == bet_team:
            return 'WIN'
        elif game_result['winner'] is None:
            return 'PENDING'
        else:
            return 'LOSS'

    elif bet_type == 'Over/Under':
        home_score = game_result.get('home_score')
        away_score = game_result.get('away_score')
        if home_score is None or away_score is None:
            return 'PENDING'
        try:
            actual_total = float(home_score) + float(away_score)
            raw = str(pred.get('book_line', ''))
            book_total = float(raw.split('@')[0])
        except (ValueError, TypeError):
            return 'PENDING'
        if bet_team == 'Over':
            if actual_total > book_total:   return 'WIN'
            elif actual_total == book_total: return 'PUSH'
            else:                           return 'LOSS'
        elif bet_team == 'Under':
            if actual_total < book_total:   return 'WIN'
            elif actual_total == book_total: return 'PUSH'
            else:                           return 'LOSS'

    return 'PENDING'

def evaluate_flag(row):
    runs = []
    for i in range(1, 4):
        team = row.get(f'run{i}_bet_team')
        edge = row.get(f'run{i}_edge')
        if team is not None and edge is not None:
            runs.append((i, team, edge))
    
    if len(runs) <= 1:
        return 'CONSISTENT'
    
    teams = [r[1] for r in runs]
    if len(set(teams)) > 1:
        return 'TEAM CHANGED'
    
    # Check if edge was lost or gained
    has_edge = [r[2] != 'NO BET' for r in runs]
    if True in has_edge and False in has_edge:
        if has_edge[0] and not has_edge[-1]:
            return 'EDGE LOST'
        elif not has_edge[0] and has_edge[-1]:
            return 'EDGE GAINED'
    
    return 'CONSISTENT'

def generate_excel_log(df):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill

        excel_file = 'predictions/predictions_log.xlsx'
        df.to_excel(excel_file, index=False)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        orange = PatternFill(start_color='F4B942', end_color='F4B942', fill_type='solid')

        col_names = [c.value for c in ws[1]]
        def col_idx(name):
            try:
                return col_names.index(name) + 1
            except ValueError:
                return None

        result_col   = col_idx('result')
        change_cols  = [col_idx(f'run{n}_change') for n in range(2, 10)]

        change_colors = {
            'PICK CHANGED': orange,
            'EDGE GAINED':  green,
            'EDGE LOST':    red,
        }

        for row_idx in range(2, ws.max_row + 1):
            # Color entire row by result
            if result_col:
                result = ws.cell(row=row_idx, column=result_col).value
                row_fill = green if result == 'WIN' else (red if result == 'LOSS' else None)
                if row_fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = row_fill

            # Overlay change columns with their own colors (takes priority)
            for cc in change_cols:
                if cc is None:
                    continue
                val = ws.cell(row=row_idx, column=cc).value
                if val in change_colors:
                    ws.cell(row=row_idx, column=cc).fill = change_colors[val]

        wb.save(excel_file)
    except Exception as e:
        print(f"Excel generation failed: {e}")

def archive_existing_prediction_logs():
    """
    One-time migration: archive predictions_log CSV/XLSX if they pre-date the
    dual-strategy schema (detected by absence of 'model_strategy' column).
    Moves files to predictions/archive/ with a date-stamped name, then creates
    fresh blank files containing the updated headers.
    Safe to call on every startup — skips if already migrated.
    """
    import shutil

    csv_file  = 'predictions/predictions_log.csv'
    xlsx_file = 'predictions/predictions_log.xlsx'
    archive_dir = 'predictions/archive'

    # Nothing to archive if neither file exists
    if not os.path.exists(csv_file) and not os.path.exists(xlsx_file):
        return

    # Already migrated — skip
    if os.path.exists(csv_file):
        try:
            peek = pd.read_csv(csv_file, nrows=1, dtype=str)
            if 'model_strategy' in peek.columns and 'recommended_stake' in peek.columns:
                return
        except Exception:
            pass  # unreadable file → archive it

    os.makedirs(archive_dir, exist_ok=True)
    suffix = f"_archived_{date.today().strftime('%Y%m%d')}"

    if os.path.exists(csv_file):
        dest = os.path.join(archive_dir, f'predictions_log{suffix}.csv')
        shutil.move(csv_file, dest)
        print(f"[Archive] {csv_file} → {dest}")

    if os.path.exists(xlsx_file):
        dest = os.path.join(archive_dir, f'predictions_log{suffix}.xlsx')
        shutil.move(xlsx_file, dest)
        print(f"[Archive] {xlsx_file} → {dest}")

    new_columns = [
        'date', 'game_time', 'game_num', 'home_team', 'away_team', 'bet_type', 'model_strategy',
        'run1_bet_team', 'run1_model_pct', 'run1_book_line', 'run1_edge',
        'run2_bet_team', 'run2_model_pct', 'run2_book_line', 'run2_edge', 'run2_change',
        'run3_bet_team', 'run3_model_pct', 'run3_book_line', 'run3_edge', 'run3_change',
        'run4_bet_team', 'run4_model_pct', 'run4_book_line', 'run4_edge', 'run4_change',
        'run5_bet_team', 'run5_model_pct', 'run5_book_line', 'run5_edge', 'run5_change',
        'run6_bet_team', 'run6_model_pct', 'run6_book_line', 'run6_edge', 'run6_change',
        'run7_bet_team', 'run7_model_pct', 'run7_book_line', 'run7_edge', 'run7_change',
        'run8_bet_team', 'run8_model_pct', 'run8_book_line', 'run8_edge', 'run8_change',
        'run9_bet_team', 'run9_model_pct', 'run9_book_line', 'run9_edge', 'run9_change',
        'final_run', 'final_bet_team', 'final_model_pct', 'final_book_line', 'final_edge',
        'raw_model_pct', 'home_platoon_factor', 'away_platoon_factor', 'platoon_confirmed',
        'recommended_stake',
        'actual_home_score', 'actual_away_score', 'winner', 'result', 'ou_direction',
    ]

    pd.DataFrame(columns=new_columns).to_csv(csv_file, index=False)
    print(f"[Archive] Created fresh {csv_file} with updated headers")

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(new_columns)
        wb.save(xlsx_file)
        print(f"[Archive] Created fresh {xlsx_file} with updated headers")
    except Exception as e:
        print(f"[Archive] Fresh XLSX creation failed: {e}")


def calculate_rolling_lambda(team,rolling):
    if team not in rolling:
        return None
    
    lg_avg = league_avg[league_avg['year'] == 2026]['lg_avg_runs'].values[0]
    
    hit_7 = rolling[team]['hit_7']
    hit_15 = rolling[team]['hit_15']
    pitch_7 = rolling[team]['pitch_7']
    pitch_15 = rolling[team]['pitch_15']

    off_rating_7 = hit_7 / lg_avg
    off_rating_15 = hit_15 / lg_avg
    pitch_rating_7 = lg_avg / pitch_7
    pitch_rating_15 = lg_avg / pitch_15

    off_rating = (off_rating_7 * 0.60) + (off_rating_15 * 0.40)
    pitch_rating = (pitch_rating_7 * 0.60) + (pitch_rating_15 * 0.40)

    off_rating = max(0.60, min(1.60, off_rating))
    pitch_rating = max(0.60, min(1.60, pitch_rating))

    return {
        'off_rating': off_rating,
        'pitch_rating': pitch_rating
    }

def get_pitcher_stats(pitcher_name, season=2026):
    try:
        search = statsapi.lookup_player(pitcher_name)
        if not search:
            return None
        pitcher_id = search[0]['id']
        stats = statsapi.player_stat_data(pitcher_id, group='pitching', type='season', sportId=1)
        if not stats['stats']:
            return None
        s = stats['stats'][0]['stats']
        innings = float(s.get('inningsPitched',0))
        runs = int(s.get('runs',0))
        ra9 = round((runs / innings * 9), 2) if innings > 0 else None
        return {
            'name': pitcher_name,
            'era': float(s.get('era', 0)),
            'innings': innings,
            'runs_allowed': runs,
            'games_started': int(s.get('gamesStarted', 0)),
            'ra9': ra9
        }
    except Exception as e:
        return None

def get_todays_starters(game_date=None):
    if game_date is None:
        game_date = date.today().strftime('%Y-%m-%d')
    schedule = statsapi.schedule(date=game_date)
    starters = {}
    for game in schedule:
        home = game['home_name']
        away = game['away_name']
        home_pitcher = game.get('home_probable_pitcher', '')
        away_pitcher = game.get('away_probable_pitcher', '')
        if home in NAME_TO_FG:
            home_fg = NAME_TO_FG[home]
            starters[home_fg] = get_pitcher_stats(home_pitcher) if home_pitcher else None
        if away in NAME_TO_FG:
            away_fg = NAME_TO_FG[away]
            starters[away_fg] = get_pitcher_stats(away_pitcher) if away_pitcher else None
    return starters

def get_total_line(game):
    best_total = None
    best_over_price = None
    best_under_price = None
    best_total_book = None
    
    for bookmaker in game['bookmakers']:
        for market in bookmaker['markets']:
            if market['key'] == 'totals':
                for outcome in market['outcomes']:
                    if outcome['name'] == 'Over':
                        if best_over_price is None or outcome['price'] > best_over_price:
                            best_over_price = outcome['price']
                            best_total = outcome['point']
                            best_total_book = bookmaker['title']
                    elif outcome['name'] == 'Under':
                        if best_under_price is None or outcome['price'] > best_under_price:
                            best_under_price = outcome['price']
    
    if best_total is None:
        return None
    
    return {
        'total': best_total,
        'over_price': best_over_price,
        'under_price': best_under_price,
        'book': best_total_book
    }

def calculate_league_averages(all_woba, all_fip, team_xwoba):
    # Offensive averages
    lg_woba  = np.mean([v['woba'] for v in all_woba.values()])
    lg_xwoba = np.mean(list(team_xwoba.values())) if team_xwoba else 0.320
    babips   = [v['babip'] for v in all_woba.values() if v.get('babip') is not None]
    lg_babip_off = np.mean(babips) if babips else 0.300

    # Pitching averages
    lg_fip      = np.mean([v['fip']   for v in all_fip.values()])
    lg_xfip     = np.mean([v['xfip']  for v in all_fip.values()])
    lg_k_pct_pit = np.mean([v['k_pct'] for v in all_fip.values()])
    lg_bb_pct_pit = np.mean([v['bb_pct'] for v in all_fip.values()])

    return {
        'lg_woba':      lg_woba,
        'lg_xwoba':     lg_xwoba,
        'lg_babip_off': lg_babip_off,
        'lg_fip':       lg_fip,
        'lg_xfip':      lg_xfip,
        'lg_k_pct_pit': lg_k_pct_pit,
        'lg_bb_pct_pit': lg_bb_pct_pit,
    }

def calculate_team_ratings(team, rolling, all_woba, all_fip, team_xwoba, lg_avgs, lg_avg_runs, injury_adj=None):
    # Rolling averages — 7-day weighted 70%, 15-day 30% (backtest optimized)
    hit_7  = rolling[team]['hit_7']
    hit_15 = rolling[team]['hit_15']
    pitch_7  = rolling[team]['pitch_7']
    pitch_15 = rolling[team]['pitch_15']
    rolling_off = (hit_7 * 0.70) + (hit_15 * 0.30)
    rolling_pit = (pitch_7 * 0.70) + (pitch_15 * 0.30)

    rolling_off_rating = rolling_off / lg_avg_runs
    rolling_pit_rating = lg_avg_runs / rolling_pit if rolling_pit > 0 else 1.0

    # Offensive ratings
    woba_rating  = all_woba[team]['woba'] / lg_avgs['lg_woba'] if team in all_woba else 1.0
    xwoba_rating = team_xwoba[team] / lg_avgs['lg_xwoba'] if team in team_xwoba and lg_avgs['lg_xwoba'] > 0 else 1.0
    babip_val    = all_woba[team].get('babip') if team in all_woba else None
    babip_rating = (babip_val / lg_avgs['lg_babip_off']
                    if babip_val is not None and lg_avgs['lg_babip_off'] > 0 else 1.0)

    # Pitching ratings — lower FIP/xFIP is better, so invert
    fip_rating   = lg_avgs['lg_fip']  / all_fip[team]['fip']  if team in all_fip and all_fip[team]['fip']  > 0 else 1.0
    xfip_rating  = lg_avgs['lg_xfip'] / all_fip[team]['xfip'] if team in all_fip and all_fip[team]['xfip'] > 0 else 1.0
    k_pit_rating  = all_fip[team]['k_pct'] / lg_avgs['lg_k_pct_pit'] if team in all_fip and lg_avgs['lg_k_pct_pit'] > 0 else 1.0
    bb_pit_rating = lg_avgs['lg_bb_pct_pit'] / all_fip[team]['bb_pct'] if team in all_fip and all_fip[team]['bb_pct'] > 0 else 1.0

    off_rating = (
        rolling_off_rating * OFFENSE_WEIGHTS['rolling'] +
        woba_rating        * OFFENSE_WEIGHTS['woba']   +
        xwoba_rating       * OFFENSE_WEIGHTS['xwoba']  +
        babip_rating       * OFFENSE_WEIGHTS['babip']
    )

    pit_rating = (
        rolling_pit_rating * PITCHING_WEIGHTS['rolling'] +
        fip_rating         * PITCHING_WEIGHTS['fip']    +
        xfip_rating        * PITCHING_WEIGHTS['xfip']   +
        k_pit_rating       * PITCHING_WEIGHTS['k_pct']  +
        bb_pit_rating      * PITCHING_WEIGHTS['bb_pct']
    )

    # Apply IL injury adjustments (tapered — only active for first 14 days)
    if injury_adj and team in injury_adj:
        adj = injury_adj[team]
        off_rating -= adj['off_adj']
        pit_rating -= adj['pit_adj']

    # Caps from backtest: low=0.65, high=1.50
    off_rating = max(0.65, min(1.50, off_rating))
    pit_rating = max(0.65, min(1.50, pit_rating))

    return off_rating, pit_rating

def get_cached_stats(team_ids, game_date=None):
    cache_file = 'cache/stats_cache.json'
    
    if os.path.exists(cache_file):
        modified_time = os.path.getmtime(cache_file)
        from datetime import datetime
        age_hours = (datetime.now().timestamp() - modified_time) / 3600
        if age_hours < 4:
            print("Loading stats from cache...")
            with open(cache_file, 'r') as f:
                return json.load(f)
    
    print("Fetching fresh stats...")
    stats = {
        'all_woba': get_all_team_woba(team_ids),
        'all_fip': get_all_team_fip(team_ids),
        'team_xwoba': get_team_xwoba(team_ids),
        'rolling': get_all_rolling_averages(team_ids),
        'starters': get_todays_starters(game_date=game_date)
    }
    
    with open(cache_file, 'w') as f:
        json.dump(convert_to_serializable(stats), f)
    
    return stats
    
def convert_to_serializable(obj):
    if isinstance(obj, np.float64):
        return float(obj)
    if isinstance(obj, np.int64):
        return int(obj)
    if isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [convert_to_serializable(i) for i in obj]
    return obj

def _get_rolling_batter_woba(player_id, min_pa=100, season=2026):
    """MLB API game log → rolling woba over last min_pa plate appearances."""
    try:
        resp = requests.get(
            f'https://statsapi.mlb.com/api/v1/people/{player_id}/stats',
            params={'stats': 'gameLog', 'group': 'hitting', 'season': season, 'sportId': 1},
            timeout=8
        ).json()
        splits = resp.get('stats', [{}])[0].get('splits', [])
        bb = hbp = singles = doubles = triples = hr = ab = ibb = sf = pa = 0
        for split in reversed(splits):
            s     = split.get('stat', {})
            g_pa  = int(s.get('plateAppearances', 0))
            g_ab  = float(s.get('atBats', 0))
            g_h   = float(s.get('hits', 0))
            g_d   = float(s.get('doubles', 0))
            g_t   = float(s.get('triples', 0))
            g_hr  = float(s.get('homeRuns', 0))
            g_bb  = float(s.get('baseOnBalls', 0))
            g_hbp = float(s.get('hitByPitch', 0))
            g_ibb = float(s.get('intentionalWalks', 0))
            g_sf  = float(s.get('sacFlies', 0))
            pa      += g_pa
            ab      += g_ab
            bb      += g_bb
            hbp     += g_hbp
            singles += g_h - g_d - g_t - g_hr
            doubles += g_d
            triples += g_t
            hr      += g_hr
            ibb     += g_ibb
            sf      += g_sf
            if pa >= min_pa:
                break
        if pa < 30:
            return None
        denom = ab + bb - ibb + sf + hbp
        if denom <= 0:
            return None
        return round((
            WOBA_WEIGHTS['wBB']  * (bb - ibb) +
            WOBA_WEIGHTS['wHBP'] * hbp +
            WOBA_WEIGHTS['w1B']  * singles +
            WOBA_WEIGHTS['w2B']  * doubles +
            WOBA_WEIGHTS['w3B']  * triples +
            WOBA_WEIGHTS['wHR']  * hr
        ) / denom, 3)
    except Exception:
        return None

def _get_pitcher_season_xfip(player_id, lg_hr_fb=0.115, season=2026, min_ip=10):
    """MLB API individual pitcher season stats → xFIP."""
    try:
        resp = statsapi.player_stat_data(player_id, group='pitching', type='season', sportId=1)
        if not resp or not resp.get('stats'):
            return None
        s  = resp['stats'][0]['stats']
        ip = ip_to_float(s.get('inningsPitched', '0'))
        if ip < min_ip:
            return None
        hr       = float(s.get('homeRuns', 0))
        bb       = float(s.get('baseOnBalls', 0))
        hbp      = float(s.get('hitBatsmen', 0))
        k        = float(s.get('strikeOuts', 0))
        air_outs = float(s.get('airOuts', 0))
        expected_hr = (air_outs + hr) * lg_hr_fb
        return round(((13 * expected_hr) + (3 * (bb + hbp)) - (2 * k)) / ip + FIP_CONSTANT, 3)
    except Exception:
        return None

def build_live_player_snapshot(lineup_data, splits_cache, lg_avgs, lg_hr_fb=0.115, season=2026):
    """
    Build a live player snapshot for the Log5 layer from current-season data.

    Batter woba:      rolling last 100 PA (MLB API game log, per player)
    Batter xwoba:     season xwoba from Baseball Savant leaderboard (one call)
    Pitcher xFIP:     computed from MLB API individual season stats (per pitcher)
    Pitcher xwoba_pit: season xwoba allowed from Baseball Savant pitcher leaderboard (one call)

    Cached for 4 hours. Returns dict matching player_snapshot.json schema.
    """
    cache_file = 'cache/live_snapshot_cache.json'
    if os.path.exists(cache_file):
        age_hours = (datetime.now().timestamp() - os.path.getmtime(cache_file)) / 3600
        if age_hours < 4:
            with open(cache_file, 'r') as f:
                data = json.load(f)
            n_bat = data.get('n_batters', len(data.get('batters', {})))
            n_pit = data.get('n_pitchers', len(data.get('pitchers', {})))
            print(f"[LiveSnapshot] Cache hit ({age_hours:.1f}h old) — {n_bat} batters, {n_pit} pitchers")
            return data

    print("[LiveSnapshot] Building from live data...")

    batter_ids  = set()
    pitcher_ids = set()
    for gd in lineup_data.values():
        batter_ids.update(gd.get('home_lineup', []))
        batter_ids.update(gd.get('away_lineup', []))
        if gd.get('home_pitcher_id'): pitcher_ids.add(gd['home_pitcher_id'])
        if gd.get('away_pitcher_id'): pitcher_ids.add(gd['away_pitcher_id'])

    hdrs = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:138.0) Gecko/20100101 Firefox/138.0'}

    savant_bat_xwoba = {}
    try:
        url = (f"https://baseballsavant.mlb.com/leaderboard/expected_statistics"
               f"?type=batter&year={season}&position=&team=&min=1&csv=true")
        df_bat = pd.read_csv(StringIO(requests.get(url, headers=hdrs, timeout=15).text))
        savant_bat_xwoba = dict(zip(df_bat['player_id'].astype(int), df_bat['est_woba'].astype(float)))
        print(f"[LiveSnapshot] Savant batters: {len(savant_bat_xwoba)} players")
    except Exception as e:
        print(f"[LiveSnapshot] Savant batter fetch failed: {e}")

    savant_pit_xwoba = {}
    try:
        url = (f"https://baseballsavant.mlb.com/leaderboard/expected_statistics"
               f"?type=pitcher&year={season}&position=&team=&min=1&csv=true")
        df_pit = pd.read_csv(StringIO(requests.get(url, headers=hdrs, timeout=15).text))
        savant_pit_xwoba = dict(zip(df_pit['player_id'].astype(int), df_pit['est_woba'].astype(float)))
        print(f"[LiveSnapshot] Savant pitchers: {len(savant_pit_xwoba)} players")
    except Exception as e:
        print(f"[LiveSnapshot] Savant pitcher fetch failed: {e}")

    batters   = {}
    n_rolling = 0
    for pid in batter_ids:
        rolling_woba = _get_rolling_batter_woba(pid, min_pa=100, season=season)
        if rolling_woba is not None:
            n_rolling += 1

        bat_info   = splits_cache.get(f"bat_{pid}", {})
        splits     = bat_info.get('splits') or {}
        xwoba      = savant_bat_xwoba.get(pid)

        entry = {}
        for hand in ('L', 'R'):
            split_key = 'vs_lhp' if hand == 'L' else 'vs_rhp'
            # Platoon split woba from MLB API statSplits (min 30 PA); fall back to rolling overall
            w = splits.get(split_key) or rolling_woba
            if w is not None or xwoba is not None:
                entry[hand] = {'woba': w, 'xwoba': xwoba}

        if entry:
            batters[str(pid)] = entry

    print(f"[LiveSnapshot] Batters built: {len(batters)} ({n_rolling} with rolling 100-PA woba)")

    pitchers = {}
    for pid in pitcher_ids:
        xfip      = _get_pitcher_season_xfip(pid, lg_hr_fb=lg_hr_fb, season=season)
        xwoba_pit = savant_pit_xwoba.get(pid)
        if xfip is not None or xwoba_pit is not None:
            pitchers[str(pid)] = {'xfip': xfip, 'xwoba_pit': xwoba_pit}

    print(f"[LiveSnapshot] Pitchers built: {len(pitchers)}")

    lg_pit_xwoba = (float(np.mean(list(savant_pit_xwoba.values())))
                    if savant_pit_xwoba else 0.315)

    snapshot = {
        'batters':   batters,
        'pitchers':  pitchers,
        'lg_avgs': {
            'L':   {'woba': lg_avgs['lg_woba'], 'xwoba': lg_avgs['lg_xwoba']},
            'R':   {'woba': lg_avgs['lg_woba'], 'xwoba': lg_avgs['lg_xwoba']},
            'pit': {'xfip': lg_avgs['lg_xfip'], 'xwoba_pit': lg_pit_xwoba},
        },
        'as_of':     datetime.now().strftime('%Y-%m-%d %H:%M'),
        'n_batters': len(batters),
        'n_pitchers': len(pitchers),
    }

    with open(cache_file, 'w') as f:
        json.dump(convert_to_serializable(snapshot), f)

    return snapshot

def update_google_sheets(df):
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds_path = os.getenv('GOOGLE_CREDENTIALS_PATH', 'cache/google_credentials.json')
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open('Baseball Model Predictions').sheet1

        # Clear cell values only from row 2 downward — preserves Row 1 headers,
        # custom background colors, borders, data validation dropdowns, and
        # conditional formatting rules (values.clear does not touch formatting).
        sheet.batch_clear(["A2:BO2000"])

        headers = df.columns.tolist()
        rows = [[str(x) if x is not None else '' for x in row.tolist()] for _, row in df.iterrows()]
        # Write headers to row 1 (in case schema changed) + data from row 2 onward
        sheet.update(range_name='A1', values=[headers] + rows)

        # Apply alternating row colors by date and game — no formula column needed
        sheet_id  = sheet.id
        n_cols    = len(df.columns)
        WHITE     = {'red': 1.0, 'green': 1.0, 'blue': 1.0}

        # Two date colors alternating — blue one day, green the next, repeat
        DATE_COLORS = [
            ({'red': 0.812, 'green': 0.886, 'blue': 0.953}, WHITE),  # light blue / no color
            ({'red': 0.812, 'green': 0.953, 'blue': 0.816}, WHITE),  # light green / no color
        ]

        # Build date order (DataFrame is newest-first)
        date_order = {}
        for d in df['date']:
            if d not in date_order:
                date_order[d] = len(date_order)

        row_within_date = {}
        color_requests  = [{
            'repeatCell': {
                'range': {'sheetId': sheet_id, 'startRowIndex': 0,
                          'endRowIndex': len(df) + 2, 'startColumnIndex': 0,
                          'endColumnIndex': n_cols},
                'cell': {'userEnteredFormat': {'backgroundColor': WHITE}},
                'fields': 'userEnteredFormat.backgroundColor',
            }
        }]

        for sheet_row, (_, row) in enumerate(df.iterrows(), start=2):
            date_val = str(row.get('date', ''))
            row_within_date[date_val] = row_within_date.get(date_val, 0) + 1
            date_idx = date_order.get(date_val, 0)
            row_idx  = row_within_date[date_val]
            color    = DATE_COLORS[date_idx % len(DATE_COLORS)][(row_idx - 1) % 2]

            color_requests.append({
                'repeatCell': {
                    'range': {'sheetId': sheet_id,
                              'startRowIndex': sheet_row - 1, 'endRowIndex': sheet_row,
                              'startColumnIndex': 0, 'endColumnIndex': n_cols},
                    'cell': {'userEnteredFormat': {'backgroundColor': color}},
                    'fields': 'userEnteredFormat.backgroundColor',
                }
            })

        sheet.spreadsheet.batch_update({'requests': color_requests})

        _tee._real.write("Google Sheets updated successfully\n")
    except Exception as e:
        import traceback
        print(f"Google Sheets update failed: {e}")
        traceback.print_exc()

def _roi_for_bets(bets_df):
    profit = 0
    wagered = 0
    for _, row in bets_df.iterrows():
        if row['result'] not in ('WIN', 'LOSS'):
            continue
        raw      = str(row.get('book_line', ''))
        bet_type = str(row.get('bet_type', ''))
        try:
            if '@' in raw:
                # O/U format "8.5@-110" — odds are after the @
                line = float(raw.split('@')[1])
            elif bet_type == 'Over/Under':
                # Old O/U row — no odds stored, skip to avoid corruption
                continue
            else:
                line = float(raw)
        except (ValueError, TypeError):
            continue
        wagered += 100
        if row['result'] == 'WIN':
            profit += line if line > 0 else 100 * (100 / abs(line))
        else:
            profit -= 100
    return (profit / wagered * 100) if wagered > 0 else 0

def send_email_report(body, run_label):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import html as _html

    sender    = os.getenv('EMAIL_SENDER')
    password  = os.getenv('EMAIL_APP_PASSWORD')
    recipient = os.getenv('EMAIL_RECIPIENT')

    if not all([sender, password, recipient]):
        print("Email not configured — skipping.")
        return

    try:
        subject = f"Baseball Model - {date.today().strftime('%#m/%#d')}"

        escaped = _html.escape(body)
        html_body = f"""\
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
  body {{ margin: 0; padding: 0; background: #ffffff; }}
  .wrap {{ max-width: 900px; margin: 0 auto; padding: 12px; }}
  pre {{
    font-family: 'Courier New', Courier, monospace;
    font-size: 14px;
    line-height: 1.5;
    white-space: pre;
    overflow-x: auto;
    background: #f9f9f9;
    border: 1px solid #e0e0e0;
    border-radius: 4px;
    padding: 12px;
    margin: 0;
  }}
  @media only screen and (max-width: 600px) {{
    .wrap {{ padding: 6px; }}
    pre {{
      font-size: 11px;
      line-height: 1.4;
      white-space: pre-wrap;
      word-break: break-word;
    }}
  }}
</style>
</head>
<body>
<div class="wrap"><pre>{escaped}</pre></div>
</body>
</html>"""

        msg = MIMEMultipart('alternative')
        msg['From']    = f'Betting Model <{sender}>'
        msg['To']      = recipient
        msg['Subject'] = subject
        msg.attach(MIMEText(body,      'plain', 'utf-8'))
        msg.attach(MIMEText(html_body, 'html',  'utf-8'))

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender, password)
            server.sendmail(sender, recipient, msg.as_string())

        print(f"Email sent to {recipient}")
    except Exception as e:
        print(f"Email failed: {e}")

MODEL_V2_START = '2026-05-24'   # date all major changes went live

def _edge_tier(e):
    try:
        pct = float(str(e).replace('%', '').replace('+', ''))
        if pct >= 15:   return '15%+'
        if pct >= 10:   return '10-15%'
        if pct >= 7:    return '7-10%'
        return None
    except Exception:
        return None

def _ou_edge_tier(e):
    try:
        runs = abs(float(str(e).replace('runs', '').replace('+', '').strip()))
        if runs >= 3.5:  return '3.5+ runs'
        if runs >= 2.5:  return '2.5-3.5 runs'
        if runs >= 1.5:  return '1.5-2.5 runs'
        return None
    except Exception:
        return None

def _print_period_stats(settled, pending_count, show_roi=True):
    if settled.empty:
        print(f"  No settled bets yet. ({pending_count} pending)")
        return

    w  = (settled['result'] == 'WIN').sum()
    l  = (settled['result'] == 'LOSS').sum()
    p  = (settled['result'] == 'PUSH').sum()
    wr = w / (w + l) if (w + l) > 0 else 0
    print(f"  Overall:    {w}W - {l}L - {p}P  |  Win Rate: {wr:.1%}  |  Pending: {pending_count}")
    print()

    for bet_type, label, tiers, tier_fn in [
        ('Moneyline',  'Moneyline ', ['7-10%', '10-15%', '15%+'], _edge_tier),
    ]:
        sub = settled[settled['bet_type'] == bet_type]
        if sub.empty:
            continue
        sw  = (sub['result'] == 'WIN').sum()
        sl  = (sub['result'] == 'LOSS').sum()
        sp  = (sub['result'] == 'PUSH').sum()
        swr = sw / (sw + sl) if (sw + sl) > 0 else 0
        if show_roi:
            roi = _roi_for_bets(sub)
            print(f"  {label}:  {sw}W - {sl}L - {sp}P  |  Win Rate: {swr:.1%}  |  ROI: {roi:+.1f}%")
        else:
            print(f"  {label}:  {sw}W - {sl}L - {sp}P  |  Win Rate: {swr:.1%}")
        print()
        sub = sub.copy()
        sub['tier'] = sub['edge'].apply(tier_fn)
        tier_group  = sub[sub['tier'].notna()].groupby('tier')
        if not tier_group.groups:
            continue
        print(f"  Edge tiers ({label.strip()}):")
        for tier in tiers:
            if tier in tier_group.groups:
                g   = tier_group.get_group(tier)
                tw  = (g['result'] == 'WIN').sum()
                tl  = (g['result'] == 'LOSS').sum()
                twr = tw / (tw + tl) if (tw + tl) > 0 else 0
                print(f"    {tier:15s}: {tw}W - {tl}L  |  {twr:.1%}")
        print()

def print_accuracy_report():
    log_file = 'predictions/predictions_log.csv'
    if not os.path.exists(log_file):
        return

    df = pd.read_csv(log_file, dtype=str)

    # Build one analysis row per prediction using final_bet_team
    rows = []
    for _, row in df.iterrows():
        bet_team  = row.get('final_bet_team')
        edge      = row.get('final_edge')
        book_line = row.get('final_book_line')
        if not _is_active_bet(bet_team):
            for run in range(1, 6):
                t = row.get(f'run{run}_bet_team')
                if _is_active_bet(t):
                    bet_team  = t
                    edge      = row.get(f'run{run}_edge')
                    book_line = row.get(f'run{run}_book_line')
                    break
        rows.append({
            'date':      row.get('date', ''),
            'bet_type':  row.get('bet_type', ''),
            'bet_team':  bet_team,
            'edge':      edge,
            'book_line': book_line,
            'result':    row.get('result'),
            'ou_direction': row.get('ou_direction', ''),
            'platoon_confirmed': row.get('platoon_confirmed', 'No'),
            'final_model_pct':  row.get('final_model_pct', ''),
            'raw_model_pct':    row.get('raw_model_pct', ''),
        })

    if not rows:
        print("\n=== ACCURACY REPORT ===\n  No predictions logged yet.\n")
        return

    analysis = pd.DataFrame(rows)
    bets     = analysis[analysis['bet_team'].apply(_is_active_bet)]
    settled  = bets[bets['result'].isin(['WIN', 'LOSS', 'PUSH'])]
    pending  = bets[bets['result'] == 'PENDING']
    voided   = bets[bets['result'] == 'VOID']

    current         = settled[settled['date'] >= MODEL_V2_START]
    current_pending = pending[pending['date'] >= MODEL_V2_START]

    print("\n=== ACCURACY REPORT ===\n")

    if len(voided) > 0:
        print(f"  ({len(voided)} voided — postponed/cancelled games excluded from all stats)\n")

    print(f"── Record (since {MODEL_V2_START}) ─────────────────────────")
    _print_period_stats(current, len(current_pending), show_roi=True)

    # ── Run total accuracy — model projection vs actual runs ──────────────────
    ou_all = analysis[analysis['bet_type'] == 'Over/Under'].copy()
    ou_with_dir = ou_all[ou_all['ou_direction'].apply(
        lambda x: bool(x) and str(x) not in ('', 'nan', 'None')
    )]

    if len(ou_with_dir) >= 5:
        print()
        print("── Run total accuracy (model projection vs actual) ──────")

        def parse_dir(d):
            try:
                parts = str(d).split()
                label = parts[0]
                err   = float(parts[1])
                return label, err
            except Exception:
                return None, None

        labels = ou_with_dir['ou_direction'].apply(lambda d: parse_dir(d)[0])
        errors = ou_with_dir['ou_direction'].apply(lambda d: parse_dir(d)[1]).dropna()

        n_high  = (labels == 'HIGH').sum()
        n_low   = (labels == 'LOW').sum()
        n_close = (labels == 'CLOSE').sum()
        n_total = len(ou_with_dir)

        avg_err  = errors.mean()
        avg_abs  = errors.abs().mean()

        print(f"  Games tracked: {n_total}  |  Avg error: {avg_err:+.2f} runs  |  Avg abs error: {avg_abs:.2f} runs")
        print(f"  HIGH (too high):  {n_high:3d}  ({n_high/n_total:.0%})")
        print(f"  LOW  (too low):   {n_low:3d}  ({n_low/n_total:.0%})")
        print(f"  CLOSE (±1 run):   {n_close:3d}  ({n_close/n_total:.0%})")

        if avg_err > 0.5:
            print(f"  >> Model is systematically HIGH by {avg_err:+.2f} runs on average")
        elif avg_err < -0.5:
            print(f"  >> Model is systematically LOW by {avg_err:+.2f} runs on average")
    print()

# ── Step 1: Archive old prediction logs if they pre-date the dual-strategy schema ──
archive_existing_prediction_logs()

# Determine run window and target date for this execution
_run_num, _run_window, _target_date = _compute_run_context()
_target_date_str = _target_date.strftime('%Y-%m-%d')
print(f"[Run context] Window: {_run_window} | Target date: {_target_date_str}")

odds_data = get_cached_odds()

upcoming = get_upcoming_games(odds_data, target_date=_target_date)

game_statuses = get_todays_game_statuses(target_date=_target_date)

in_progress_games = []
completed_games = []

team_ids = get_all_team_ids()

_log5_bundle, _player_snapshot = load_log5_assets()

from datetime import date
first_game_date = _target_date_str
cached_stats = get_cached_stats(team_ids, game_date=first_game_date)

all_woba = cached_stats['all_woba']
all_fip = cached_stats['all_fip']
team_xwoba = cached_stats['team_xwoba']
rolling = cached_stats['rolling']
todays_starters = cached_stats.get('starters', {})
lg_avgs = calculate_league_averages(all_woba, all_fip, team_xwoba)
injury_adj    = get_injury_adjustments(team_ids, all_woba, all_fip, lg_avgs)
lineups       = get_confirmed_lineups(first_game_date)
platoon_splits = get_all_platoon_splits(lineups) if lineups else {}

# Build live player snapshot for Log5 layer — replaces static player_snapshot.json
# Uses rolling 100 PA woba (MLB game log per player) + season xwoba (Savant leaderboard)
_live_snapshot = None
if lineups and _log5_bundle:
    _lg_hr_fb = next((v.get('lg_hr_fb', 0.115) for v in all_fip.values() if v.get('lg_hr_fb')), 0.115)
    _live_snapshot = build_live_player_snapshot(lineups, platoon_splits, lg_avgs, lg_hr_fb=_lg_hr_fb)

for key, status in game_statuses.items():
    parts = key.split('_')
    away_fg = parts[0]
    home_fg = parts[1]
    
    away_name = next((k for k, v in NAME_TO_FG.items() if v == away_fg), away_fg)
    home_name = next((k for k, v in NAME_TO_FG.items() if v == home_fg), home_fg)
    
    if status['status'] == 'In Progress':
        in_progress_games.append(f"⚾ {away_name} @ {home_name} | {status['away_score']} - {status['home_score']}")
    elif status['status'] == 'Final':
        completed_games.append(f"✔️  {away_name} @ {home_name} | Final: {status['away_score']} - {status['home_score']}")

# Reset capture buffer here so startup noise (API key, cache messages) is excluded from email
_tee._buf.truncate(0)
_tee._buf.seek(0)

print(f"\nUpcoming games today: {len(upcoming)}")

# Effective bankroll = total bankroll minus capital committed to currently open/un-settled bets
effective_bankroll = BANKROLL
_eff_log = 'predictions/predictions_log.csv'
if os.path.exists(_eff_log):
    try:
        _eff_df = pd.read_csv(_eff_log, dtype=str)
        if 'recommended_stake' in _eff_df.columns and 'result' in _eff_df.columns:
            _open = _eff_df[_eff_df['result'] == 'PENDING']
            _committed = pd.to_numeric(_open['recommended_stake'], errors='coerce').fillna(0).sum()
            effective_bankroll = max(0.0, BANKROLL - _committed)
    except Exception:
        pass

print(f"\n=== TODAY'S EDGES ===")
print(f"Bankroll: ${BANKROLL:,.0f} | Effective: ${effective_bankroll:,.0f} | Kelly: ½ ({KELLY_FRACTION}x) | Cap: {MAX_BANKROLL_EXPOSURE:.0%}/bet | Sniper >4.5% / Enforcer >4.0%\n")

edge_games = []
no_edge_games = []
skipped_games = []
summary_lines = []

# Load today's prior predictions for change detection (run N-1 vs current run)
prior_picks_today = {}  # {(home_fg, away_fg, game_num_str): last_bet_team}
_prior_log = 'predictions/predictions_log.csv'
if os.path.exists(_prior_log):
    try:
        _prior_df = pd.read_csv(_prior_log, dtype=str)
        _today_ml = _prior_df[
            (_prior_df['date'] == first_game_date) &
            (_prior_df['bet_type'] == 'Moneyline')
        ]
        for _, _row in _today_ml.iterrows():
            _gn = str(_row.get('game_num', '1') or '1')
            if _gn in ('', 'nan', 'None'):
                _gn = '1'
            _key = (_row['home_team'], _row['away_team'], _gn)
            for _r in [3, 2, 1]:
                _bt = _row.get(f'run{_r}_bet_team')
                if pd.notna(_bt) and str(_bt) not in ('', 'nan', 'None'):
                    prior_picks_today[_key] = str(_bt)
                    break
    except Exception:
        pass

n_pick_changed = 0
n_new_edge     = 0
n_edge_lost    = 0

# Game Loop
upcoming = sorted(upcoming, key=lambda x: x['commence_time'])
pair_count      = {}
active_game_keys = set()  # all games visible in today's odds feed
for game in upcoming:
    home_name = game['home_team']
    away_name = game['away_team']

    if home_name not in NAME_TO_FG or away_name not in NAME_TO_FG:
        continue

    home_fg = NAME_TO_FG[home_name]
    away_fg = NAME_TO_FG[away_name]

    base_mlb_key = f"{away_fg}_{home_fg}"
    pair_count[base_mlb_key] = pair_count.get(base_mlb_key, 0) + 1
    game_num = pair_count[base_mlb_key]
    mlb_key  = base_mlb_key if game_num == 1 else f"{base_mlb_key}_G2"
    active_game_keys.add(mlb_key)

    mlb_status = game_statuses.get(mlb_key, {}).get('status', '')
    if mlb_status in ('Postponed', 'Cancelled'):
        print(f"  {away_fg} @ {home_fg} — {mlb_status}, skipping")
        continue

    if home_fg not in rolling or away_fg not in rolling:
        continue

    if len(game['bookmakers']) == 0:
        continue

    best_home_line = None
    best_away_line = None
    best_home_book = None
    best_away_book = None

    for bookmaker in game['bookmakers']:
        for market in bookmaker['markets']:
            if market['key'] == 'h2h':
                for outcome in market['outcomes']:
                    if outcome['name'] == home_name:
                        if best_home_line is None or outcome['price'] > best_home_line:
                            best_home_line = outcome['price']
                            best_home_book = bookmaker['title']
                    elif outcome['name'] == away_name:
                        if best_away_line is None or outcome['price'] > best_away_line:
                            best_away_line = outcome['price']
                            best_away_book = bookmaker['title']

    if best_home_line is None or best_away_line is None:
        continue

    home_market_line = best_home_line
    away_market_line = best_away_line

    home_market_prob = american_to_prob(home_market_line)
    away_market_prob = american_to_prob(away_market_line)

    game_time = datetime.strptime(game['commence_time'], '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc).astimezone(MT)
    dh_label = f' (Game {game_num})' if game_num > 1 else ''
    tz_label = 'MDT' if game_time.dst() else 'MST'
    time_str = game_time.strftime(f'%a %b %d - %I:%M %p {tz_label}') + dh_label

    home_starter = todays_starters.get(home_fg)
    away_starter = todays_starters.get(away_fg)

    if home_starter is None or away_starter is None:
        missing = []
        if home_starter is None:
            missing.append(home_name)
        if away_starter is None:
            missing.append(away_name)
        skipped_games.append(f"⚠️  {away_name} @ {home_name} - {time_str} | Starter not yet announced for: {', '.join(missing)}")
        continue

    low_sample = []
    if home_starter and home_starter['innings'] < MIN_INNINGS:
        low_sample.append(f"{home_name} starter {home_starter['name']} ({home_starter['innings']} IP)")
    if away_starter and away_starter['innings'] < MIN_INNINGS:
        low_sample.append(f"{away_name} starter {away_starter['name']} ({away_starter['innings']} IP)")

    # Compute platoon factors if lineup is confirmed for this game
    platoon_data  = None
    platoon_text  = ""
    game_key_plat = (home_fg, away_fg)
    if game_key_plat in lineups and platoon_splits:
        gd = lineups[game_key_plat]
        if gd['home_lineup'] and gd['away_lineup']:
            h_factor, a_factor, plat_notes = compute_platoon_factors(
                gd['home_lineup'], gd['away_lineup'],
                gd['home_pitcher_id'], gd['away_pitcher_id'],
                platoon_splits, lg_avgs
            )
            platoon_data = (h_factor, a_factor)
            platoon_text = (
                f"\n   📐 Platoon — {home_name}: {h_factor - 1:+.1%} | "
                f"{away_name}: {a_factor - 1:+.1%}  ({plat_notes})"
            )
        else:
            platoon_text = "\n   📐 Platoon — Lineup not yet confirmed"
    else:
        platoon_text = "\n   📐 Platoon — Lineup not yet confirmed"

    # Determine whether this game has confirmed lineups
    _lineup_confirmed = (
        game_key_plat in lineups and
        bool(lineups[game_key_plat].get('home_lineup')) and
        bool(lineups[game_key_plat].get('away_lineup'))
    )
    _pre_lineup = not _lineup_confirmed

    # Always compute raw (no platoon) result for tracking purposes
    raw_result = calculate_matchup(home_fg, away_fg, 2026, rolling=rolling, starters=todays_starters,
                                    injury_adj=injury_adj, platoon_data=None)
    if raw_result is None:
        continue
    raw_home_pct, raw_away_pct = raw_result[0], raw_result[1]

    result = calculate_matchup(home_fg, away_fg, 2026, rolling=rolling, starters=todays_starters,
                                injury_adj=injury_adj, platoon_data=platoon_data)
    if result is None:
        continue

    home_win_pct, away_win_pct, avg_total, home_lambda, away_lambda = result
    total_data = get_total_line(game)

    # Use Log5 regression for win probability — confirmed lineup or projected order
    _active_snapshot = _live_snapshot or _player_snapshot
    _proj_lineup_used = False
    _log5_stats = None
    if _log5_bundle and _active_snapshot and not _pre_lineup:
        # ── Confirmed lineups: full Log5 with actual batting order ────────────
        gd = lineups.get(game_key_plat, {})
        _h_pid = gd.get('home_pitcher_id')
        _a_pid = gd.get('away_pitcher_id')
        _h_hand = (platoon_splits.get(f"pit_{_h_pid}") or {}).get('hand', 'R')
        _a_hand = (platoon_splits.get(f"pit_{_a_pid}") or {}).get('hand', 'R')
        _log5_result = compute_log5_win_prob(
            gd.get('home_lineup', []), gd.get('away_lineup', []),
            _h_pid, _a_pid, _h_hand, _a_hand,
            _log5_bundle, _active_snapshot,
        )
        if _log5_result is not None:
            home_win_pct, away_win_pct, _log5_stats = _log5_result

    elif _log5_bundle and _active_snapshot and _pre_lineup:
        # ── Pre-lineup: probable pitchers + projected batting order ──────────
        # Only runs when both probable starters have been announced. Batting
        # order is the most-common player per slot across the last 10 games.
        # Bet rec from this run is automatically compared against the later
        # confirmed-lineup run by the existing change-detection logic.
        _prob_gd = lineups.get(game_key_plat, {})
        _h_pid   = _prob_gd.get('home_pitcher_id')
        _a_pid   = _prob_gd.get('away_pitcher_id')
        if _h_pid and _a_pid:
            _h_hand   = (platoon_splits.get(f"pit_{_h_pid}") or {}).get('hand', 'R')
            _a_hand   = (platoon_splits.get(f"pit_{_a_pid}") or {}).get('hand', 'R')
            _proj_home = build_projected_lineup(home_fg)
            _proj_away = build_projected_lineup(away_fg)
            if len(_proj_home) >= 5 and len(_proj_away) >= 5:
                _log5_result = compute_log5_win_prob(
                    _proj_home, _proj_away,
                    _h_pid, _a_pid, _h_hand, _a_hand,
                    _log5_bundle, _active_snapshot,
                )
                if _log5_result is not None:
                    home_win_pct, away_win_pct, _log5_stats = _log5_result
                    _proj_lineup_used = True

    # Reflect projected-lineup status in the platoon display line
    if _proj_lineup_used:
        _prob_gd = lineups.get(game_key_plat, {})
        _h_starter = todays_starters.get(home_fg, {})
        _a_starter = todays_starters.get(away_fg, {})
        _h_pname = _h_starter.get('name', 'TBD') if _h_starter else 'TBD'
        _a_pname = _a_starter.get('name', 'TBD') if _a_starter else 'TBD'
        platoon_text = (
            f"\n   📐 Projected order (last 10 games) | "
            f"Probable: {_h_pname} vs {_a_pname} | "
            f"Platoon pending lineup confirmation"
        )

    # De-vig market probabilities (display only — edge is vs raw implied prob to account for vig)
    _dv_total    = home_market_prob + away_market_prob
    _dv_home_mkt = home_market_prob / _dv_total
    _dv_away_mkt = away_market_prob / _dv_total

    home_edge = home_win_pct - home_market_prob
    away_edge = away_win_pct - away_market_prob


    # Collect injury notes for both teams
    inj_notes = []
    for team_fg in [home_fg, away_fg]:
        if injury_adj and team_fg in injury_adj and injury_adj[team_fg]['notes']:
            inj_notes.extend(injury_adj[team_fg]['notes'])
    injury_text = ("\n   🏥 " + " | ".join(inj_notes)) if inj_notes else ""

    moneyline_edge = False
    ou_edge = False
    kelly_amt = 0.0
    _ml_tier  = ''
    game_text = f"{away_name} @ {home_name} - {time_str}"

    # Tiered edge thresholds — pick the stronger side first
    _SNIPER_MIN   = 0.045
    _ENFORCER_MIN = 0.040
    _best_is_home = home_edge >= away_edge
    _best_edge    = home_edge if _best_is_home else away_edge

    if _best_edge > _SNIPER_MIN:
        _ml_tier = 'SNIPER'
    elif _best_edge > _ENFORCER_MIN:
        _ml_tier = 'ENFORCER'

    if _ml_tier and _best_is_home:
        kelly_pct, kelly_amt = kelly_bet_size(home_win_pct, home_market_line, bankroll=effective_bankroll)
        _ev_dec = (1 + home_market_line / 100) if home_market_line > 0 else (1 + 100 / abs(home_market_line))
        _ev_val = home_win_pct * kelly_amt * (_ev_dec - 1) - (1 - home_win_pct) * kelly_amt
        _ev_sign = '+' if _ev_val >= 0 else ''
        moneyline_text = (
            f"   [ACTION: {_ml_tier} BET] 💰 Bet {home_name} | Kelly: {kelly_pct}% (${kelly_amt:.2f})\n"
            f"      💵 EV (${kelly_amt:.2f} stake on {home_name}): {_ev_sign}${_ev_val:.2f}\n"
            f"      Model: {home_win_pct:.1%} | {best_home_book}: {home_market_line} | Edge: +{home_edge:.1%}"
        )
        moneyline_edge = True
        bet_fg   = home_fg
        bet_pct  = home_win_pct
        bet_line = home_market_line
        bet_edge = home_edge
    elif _ml_tier:
        kelly_pct, kelly_amt = kelly_bet_size(away_win_pct, away_market_line, bankroll=effective_bankroll)
        _ev_dec = (1 + away_market_line / 100) if away_market_line > 0 else (1 + 100 / abs(away_market_line))
        _ev_val = away_win_pct * kelly_amt * (_ev_dec - 1) - (1 - away_win_pct) * kelly_amt
        _ev_sign = '+' if _ev_val >= 0 else ''
        moneyline_text = (
            f"   [ACTION: {_ml_tier} BET] 💰 Bet {away_name} | Kelly: {kelly_pct}% (${kelly_amt:.2f})\n"
            f"      💵 EV (${kelly_amt:.2f} stake on {away_name}): {_ev_sign}${_ev_val:.2f}\n"
            f"      Model: {away_win_pct:.1%} | {best_away_book}: {away_market_line} | Edge: +{away_edge:.1%}"
        )
        moneyline_edge = True
        bet_fg   = away_fg
        bet_pct  = away_win_pct
        bet_line = away_market_line
        bet_edge = away_edge
    else:
        _nb_model = home_win_pct if _best_is_home else away_win_pct
        _nb_book  = best_home_book if _best_is_home else best_away_book
        _nb_line  = home_market_line if _best_is_home else away_market_line
        _nb_edge  = home_edge if _best_is_home else away_edge
        moneyline_text = f"   [NO BET] | Model: {_nb_model:.1%} | {_nb_book}: {_nb_line} | Edge: {_nb_edge:+.1%}"

    if moneyline_edge:
        _sum_team     = home_fg if bet_fg == home_fg else away_fg
        _sum_book     = best_home_book if bet_fg == home_fg else best_away_book
        _sum_odds     = home_market_line if bet_fg == home_fg else away_market_line
        _sum_odds_str = f"+{_sum_odds}" if _sum_odds > 0 else str(_sum_odds)
        summary_lines.append(f"✅ {away_fg} @ {home_fg}    {_ml_tier}    Bet {_sum_team}    {_sum_book} {_sum_odds_str}    ${kelly_amt:.2f}")
    else:
        summary_lines.append(f"❌ {away_fg} @ {home_fg}    No bet  ({_nb_edge:+.1%})")

    ml_stake = kelly_amt  # 0.0 when no edge

    ou_stake = 0.0
    if total_data:
        book_total = total_data['total']
        ou_diff    = avg_total - book_total

        # Derive over/under probability from Poisson total via normal approximation
        # Total runs ~ Poisson(home_lambda + away_lambda), variance = mean
        total_lambda = home_lambda + away_lambda
        z = (book_total + 0.5 - total_lambda) / math.sqrt(max(total_lambda, 0.1))
        over_prob  = 0.5 * (1 - math.erf(z / math.sqrt(2)))
        under_prob = 1.0 - over_prob

        if OU_BETTING_ENABLED and ou_diff > 1.5:
            kelly_ou_pct, kelly_ou_amt = kelly_bet_size(over_prob, total_data['over_price'], bankroll=effective_bankroll)
            ou_stake = kelly_ou_amt
            ou_text = (
                f"   📊 Over/Under: OVER | Model: {avg_total:.1f} | "
                f"{total_data['book']}: {book_total} (Over {total_data['over_price']}) | "
                f"Edge: +{ou_diff:.1f} runs | Kelly: {kelly_ou_pct}% (${kelly_ou_amt:.2f})"
            )
            ou_edge = True

        elif OU_BETTING_ENABLED and ou_diff < -1.5:
            kelly_ou_pct, kelly_ou_amt = kelly_bet_size(under_prob, total_data['under_price'], bankroll=effective_bankroll)
            ou_stake = kelly_ou_amt
            ou_text = (
                f"   📊 Over/Under: UNDER | Model: {avg_total:.1f} | "
                f"{total_data['book']}: {book_total} (Under {total_data['under_price']}) | "
                f"Edge: {ou_diff:.1f} runs | Kelly: {kelly_ou_pct}% (${kelly_ou_amt:.2f})"
            )
            ou_edge = True

        else:
            ou_text = f"   ➖ Over/Under: No edge | Model: {avg_total:.1f} | Book: {book_total}"
    else:
        ou_text = f"   ➖ Over/Under: No line available"

    # Breakdown lines for email transparency
    mkt_model_text = (
        f"\n   📊 Mkt (de-vig): {home_name} {_dv_home_mkt:.1%} / {away_name} {_dv_away_mkt:.1%}"
        f" | Model (player-only): {home_name} {home_win_pct:.1%} / {away_name} {away_win_pct:.1%}"
    )
    if _log5_stats:
        _h_vs = _log5_stats['a_pit_hand']
        _a_vs = _log5_stats['h_pit_hand']
        stats_text = (
            f"\n   🔢 {home_name} bat vs {_h_vs}HP: wOBA {_log5_stats['h_bat_woba']:.3f} / xwOBA {_log5_stats['h_bat_xwoba']:.3f}"
            f"  |  {away_name} bat vs {_a_vs}HP: wOBA {_log5_stats['a_bat_woba']:.3f} / xwOBA {_log5_stats['a_bat_xwoba']:.3f}"
        )
    else:
        _h_woba  = all_woba.get(home_fg, {}).get('woba', 0.0)
        _a_woba  = all_woba.get(away_fg, {}).get('woba', 0.0)
        _h_xwoba = team_xwoba.get(home_fg, 0.0)
        _a_xwoba = team_xwoba.get(away_fg, 0.0)
        stats_text = (
            f"\n   🔢 {home_name}: wOBA {_h_woba:.3f} / xwOBA {_h_xwoba:.3f}"
            f"  |  {away_name}: wOBA {_a_woba:.3f} / xwOBA {_a_xwoba:.3f}"
        )

    _hs_name = home_starter['name'] if home_starter else 'TBD'
    _hs_era  = f"{home_starter['era']:.2f}" if home_starter else 'N/A'
    _hs_ip   = f"{home_starter['innings']:.1f}" if home_starter else '0.0'
    _as_name = away_starter['name'] if away_starter else 'TBD'
    _as_era  = f"{away_starter['era']:.2f}" if away_starter else 'N/A'
    _as_ip   = f"{away_starter['innings']:.1f}" if away_starter else '0.0'
    if _log5_stats:
        _hs_xfip = f" / {_log5_stats['h_pit_xfip']:.2f} xFIP"
        _as_xfip = f" / {_log5_stats['a_pit_xfip']:.2f} xFIP"
    else:
        _hs_xfip = ""
        _as_xfip = ""
    starters_text = (
        f"\n   ⚾  Starters: {_hs_name} {_hs_era} ERA{_hs_xfip} ({_hs_ip} IP)"
        f"  vs  {_as_name} {_as_era} ERA{_as_xfip} ({_as_ip} IP)"
    )

    _post_starters = ""
    if low_sample:
        _post_starters += f"\n      ⚠️  Low sample: {', '.join(low_sample)}"

    _breakdown = f"{mkt_model_text}{stats_text}{starters_text}{_post_starters}"

    # Log moneyline for all games
    if moneyline_edge:
        ml_bet_team = bet_fg
        ml_model_pct = f"{bet_pct:.1%}"
        ml_book_line = str(bet_line)
        ml_edge = f"+{bet_edge:.1%}"
    else:
        ml_bet_team = 'No Bet'
        ml_model_pct = f"{max(home_win_pct, away_win_pct):.1%}"
        ml_book_line = str(home_market_line if home_win_pct > away_win_pct else away_market_line)
        ml_edge = 'No Edge'

    # Detect changes from the most recent prior run logged today (ml_bet_team must be set first)
    change_text = ""
    _prior_bet = prior_picks_today.get((home_fg, away_fg, str(game_num)))
    if _prior_bet is not None:
        _prev_active = _is_active_bet(_prior_bet)
        if _prev_active and moneyline_edge and _prior_bet != ml_bet_team:
            n_pick_changed += 1
            change_text = f"\n   🔄 PICK CHANGED: {_prior_bet} → {ml_bet_team}"
        elif _prev_active and not moneyline_edge:
            n_edge_lost += 1
            change_text = f"\n   ⚠️  EDGE LOST: Previously bet {_prior_bet}"
        elif not _prev_active and moneyline_edge:
            n_new_edge += 1
            change_text = f"\n   🆕 NEW EDGE: {ml_bet_team} (not flagged in prior run)"

    game_time_24h = game_time.strftime('%H:%M')

    home_pf = round(platoon_data[0], 4) if platoon_data else None
    away_pf = round(platoon_data[1], 4) if platoon_data else None

    # Raw model pct for the bet team (unadjusted, for platoon comparison)
    if moneyline_edge:
        raw_ml_pct = f"{raw_home_pct:.1%}" if bet_fg == home_fg else f"{raw_away_pct:.1%}"
    else:
        raw_ml_pct = f"{max(raw_home_pct, raw_away_pct):.1%}"

    _ml_strategy_tag = f"{_ml_tier} — LISTED PITCHERS ONLY" if _ml_tier else ''
    todays_predictions.append({
        'date':               first_game_date,
        'game_time':          game_time_24h,
        'game_num':           game_num,
        'run_time':           run_time,
        'home_fg':            home_fg,
        'away_fg':            away_fg,
        'bet_type':           'Moneyline',
        'model_strategy':     _ml_strategy_tag,
        'bet_team':           ml_bet_team,
        'model_pct':          ml_model_pct,
        'raw_model_pct':      raw_ml_pct,
        'book_line':          ml_book_line,
        'edge':               ml_edge,
        'home_platoon_factor': str(home_pf) if home_pf is not None else '',
        'away_platoon_factor': str(away_pf) if away_pf is not None else '',
        'platoon_confirmed':  'Yes' if platoon_data else 'No',
        'recommended_stake':  ml_stake,
    })

    # Log O/U for all games
    if total_data:
        if ou_edge:
            ou_bet_team = 'Over' if ou_diff > 0 else 'Under'
            ou_edge_str = f"{ou_diff:+.1f} runs"
            # Store as "total@odds" so actual odds are available for ROI calculation
            ou_odds     = total_data['over_price'] if ou_diff > 0 else total_data['under_price']
            ou_book_line = f"{book_total}@{ou_odds}"
        else:
            ou_bet_team  = 'No Bet'
            ou_edge_str  = 'No Edge'
            ou_book_line = str(book_total)

        raw_ou_total = f"{raw_result[2]:.1f}"
        todays_predictions.append({
            'date':               first_game_date,
            'game_time':          game_time_24h,
            'game_num':           game_num,
            'run_time':           run_time,
            'home_fg':            home_fg,
            'away_fg':            away_fg,
            'bet_type':           'Over/Under',
            'model_strategy':     '',
            'bet_team':           ou_bet_team,
            'model_pct':          f"{avg_total:.1f}",
            'raw_model_pct':      raw_ou_total,
            'book_line':          ou_book_line,
            'edge':               ou_edge_str,
            'home_platoon_factor': str(home_pf) if home_pf is not None else '',
            'away_platoon_factor': str(away_pf) if away_pf is not None else '',
            'platoon_confirmed':  'Yes' if platoon_data else 'No',
            'recommended_stake':  ou_stake,
        })
    else:
        todays_predictions.append({
            'date':               first_game_date,
            'game_time':          game_time_24h,
            'game_num':           game_num,
            'run_time':           run_time,
            'home_fg':            home_fg,
            'away_fg':            away_fg,
            'bet_type':           'Over/Under',
            'model_strategy':     '',
            'bet_team':           'No Bet',
            'model_pct':          f"{avg_total:.1f}",
            'raw_model_pct':      f"{raw_result[2]:.1f}",
            'book_line':          'N/A',
            'edge':               'No Line',
            'home_platoon_factor': str(home_pf) if home_pf is not None else '',
            'away_platoon_factor': str(away_pf) if away_pf is not None else '',
            'platoon_confirmed':  'Yes' if platoon_data else 'No',
            'recommended_stake':  0.0,
        })

    if moneyline_edge or ou_edge:
        edge_games.append({
            'text': f"✅ {game_text}\n{moneyline_text}{injury_text}{change_text}{_breakdown}"
        })
    else:
        no_edge_games.append(f"❌ {game_text}\n{moneyline_text}{injury_text}{change_text}{_breakdown}")

# Print compact summary (mobile-friendly top section)
from datetime import date as _date
_today_str = _date.today().strftime('%a %b %d').replace(' 0', ' ')
_bets_today = [l for l in summary_lines if l.startswith('✅')]
_n_bets = len(_bets_today)
print(f"=== TODAY'S PICKS — {_today_str} | {_n_bets} bet{'s' if _n_bets != 1 else ''} ===\n")
for _sl in summary_lines:
    print(_sl)
print(f"\n{'═' * 54}\n")

# Print change summary (only on runs 2+ when prior data exists)
_change_parts = []
if n_pick_changed: _change_parts.append(f"{n_pick_changed} pick change{'s' if n_pick_changed > 1 else ''}")
if n_new_edge:     _change_parts.append(f"{n_new_edge} new edge{'s' if n_new_edge > 1 else ''}")
if n_edge_lost:    _change_parts.append(f"{n_edge_lost} edge{'s' if n_edge_lost > 1 else ''} lost")
if _change_parts:
    print(f"📋 Changes from prior run: {' | '.join(_change_parts)}\n")

print("=== DETAIL ===\n")

# Print edges
if edge_games:
    for game in edge_games:
        print(game['text'])
        print()
else:
    print("No edges found today.\n")

# Print no edge games
if no_edge_games:
    print("=== NO EDGE FOUND ===\n")
    for game in no_edge_games:
        print(game)
        print()

# Print skipped games
if skipped_games:
    print("=== AWAITING STARTERS ===\n")
    for game in skipped_games:
        print(game)
        print()


if in_progress_games:
    print("\n=== IN PROGRESS ===\n")
    for game in in_progress_games:
        print(game)
        print()

if completed_games:
    print("\n=== COMPLETED ===\n")
    for game in completed_games:
        print(game)
        print()

todays_starters = get_todays_starters()

log_predictions(todays_predictions, yesterdays_results, yesterdays_postponed, game_statuses, active_game_keys,
                run_num=_run_num, run_window=_run_window)

# Capture email body before accuracy report runs
from datetime import datetime as _dt
print_accuracy_report()
_email_body = _tee.getvalue()

_now_utc = _dt.now(timezone.utc)
if _run_window in _OVERNIGHT_WINDOWS:
    # Overnight runs target tomorrow's opening lines — always send if games were found
    if upcoming:
        send_email_report(_email_body, _run_window)
    else:
        print(f"[Email skipped] No games found for {_target_date_str}.")
else:
    # Day-of runs: only send if a game starts within 3 hours
    _email_window_hours = 3
    _has_game_soon = any(
        (_dt.strptime(g['commence_time'], '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc) - _now_utc).total_seconds()
        <= _email_window_hours * 3600
        for g in upcoming
    )
    if _has_game_soon:
        send_email_report(_email_body, _run_window)
    else:
        _next_times = sorted(
            _dt.strptime(g['commence_time'], '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc)
            for g in upcoming
        )
        if _next_times:
            _hrs_until = (_next_times[0] - _now_utc).total_seconds() / 3600
            print(f"[Email skipped] No games within {_email_window_hours}h. Next game in {_hrs_until:.1f}h.")
        else:
            print(f"[Email skipped] No upcoming games for {_target_date_str}.")
